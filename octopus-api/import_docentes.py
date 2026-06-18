#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script para importar docentes desde Excel a la BD.
Uso: cd C:\Octopus\octopus-api && python import_docentes.py
"""
import os
import sys
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from rrhh.models import Empleado, BancoNomina
from openpyxl import load_workbook

EXCEL_PATH = r"C:\Users\PC\Downloads\Base de Datos Colegio\DATOS DEL PERSONAL DOCENTE.xlsx"

def limpiar_cedula(cedula):
    """Limpia cedulas (formato V-12.345.678)."""
    if not cedula:
        return None
    cedula = str(cedula).strip().replace('.', '').replace('-', '')
    return cedula if cedula else None

def limpiar_texto(texto):
    """Limpia y normaliza texto."""
    if texto is None or texto == '':
        return ''
    return str(texto).strip()

def importar_docentes():
    """Importa docentes desde Excel a la BD."""
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active

        print(f"\n{'='*100}")
        print(f"IMPORTANDO DOCENTES DESDE EXCEL")
        print(f"{'='*100}\n")

        # Los encabezados estan en la fila 5
        HEADER_ROW = 5
        headers = {}
        for i in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=HEADER_ROW, column=i).value
            if cell_value:
                headers[cell_value.lower().strip()] = i - 1

        print(f"Encabezados encontrados: {list(headers.keys())}")
        print()

        # Validar minimo
        if 'apellido y nombre' not in headers or 'cedula' not in headers:
            print("ERROR: Columnas requeridas no encontradas")
            return 0, 1

        creados = 0
        actualizados = 0
        errores = 0
        errores_lista = []

        # Datos comienzan en fila 6
        DATA_START_ROW = HEADER_ROW + 1

        print(f"Procesando filas desde {DATA_START_ROW}...")
        print()

        for row_num in range(DATA_START_ROW, ws.max_row + 1):
            try:
                # Obtener valores
                nombre_completo = limpiar_texto(
                    ws.cell(row=row_num, column=headers['apellido y nombre'] + 1).value
                )
                cedula = limpiar_cedula(
                    ws.cell(row=row_num, column=headers['cedula'] + 1).value
                )
                correo = limpiar_texto(
                    ws.cell(row=row_num, column=headers.get('correo', -1) + 1).value
                    if 'correo' in headers else ''
                )
                telefono = limpiar_texto(
                    ws.cell(row=row_num, column=headers.get('telefono', -1) + 1).value
                    if 'telefono' in headers else ''
                )

                # Saltar vacias
                if not nombre_completo or not cedula:
                    continue

                # Parsear nombre y apellido (formato: "APELLIDO NOMBRE")
                partes = nombre_completo.split()
                if len(partes) >= 2:
                    apellido = partes[0]
                    nombre = ' '.join(partes[1:])
                else:
                    apellido = partes[0]
                    nombre = ''

                # Crear o actualizar
                empleado, creado = Empleado.objects.update_or_create(
                    cedula=cedula,
                    defaults={
                        'nombre': nombre,
                        'apellido': apellido,
                        'cargo': '',
                        'tipo_personal': 'docente',
                        'titulo': '',
                        'categoria_docente': '',
                        'nivel': '',
                        'horas_semanales': None,
                        'numero_hijos': 0,
                        'telefono': telefono,
                        'correo': correo,
                        'numero_cuenta': '',
                        'tipo_cuenta': '',
                        'sueldo_base': None,
                        'activo': True,
                    }
                )

                if creado:
                    creados += 1
                    print(f"[NUEVO] {nombre} {apellido} - {cedula}")
                else:
                    actualizados += 1
                    print(f"[ACTUALIZADO] {nombre} {apellido} - {cedula}")

            except Exception as e:
                errores += 1
                error_msg = f"Fila {row_num}: {str(e)}"
                errores_lista.append(error_msg)
                print(f"[ERROR] {error_msg}")

        # Resumen
        print(f"\n{'='*100}")
        print("RESUMEN DE IMPORTACION")
        print(f"{'='*100}")
        print(f"Creados:       {creados}")
        print(f"Actualizados:  {actualizados}")
        print(f"Errores:       {errores}")
        print(f"Total:         {creados + actualizados + errores}\n")

        if errores_lista:
            print("ERRORES:")
            for err in errores_lista[:10]:
                print(f"  - {err}")

        return creados + actualizados, errores

    except Exception as e:
        print(f"ERROR CRITICO: {e}")
        import traceback
        traceback.print_exc()
        return 0, 1

if __name__ == '__main__':
    creados, errores = importar_docentes()
    sys.exit(0 if errores == 0 else 1)
