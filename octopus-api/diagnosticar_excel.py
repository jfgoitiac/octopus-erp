#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Diagnóstico del Excel de docentes para ver su estructura."""
import sys
from openpyxl import load_workbook

EXCEL_PATH = r"C:\Users\PC\Downloads\Base de Datos Colegio\DATOS DEL PERSONAL DOCENTE.xlsx"

try:
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    print(f"\n{'='*100}")
    print(f"DIAGNOSTICO DEL ARCHIVO EXCEL")
    print(f"{'='*100}\n")
    print(f"Nombre de la hoja: {ws.title}")
    print(f"Filas totales: {ws.max_row}")
    print(f"Columnas totales: {ws.max_column}")
    print(f"Merged cells: {list(ws.merged_cells.ranges)}")

    # Buscar la fila de encabezados (la que tiene datos no vacios)
    print(f"\n Buscando fila con encabezados...")
    header_row = 1
    for row_num in range(1, min(10, ws.max_row + 1)):
        row_values = [ws.cell(row=row_num, column=col).value for col in range(1, ws.max_column + 1)]
        non_empty = sum(1 for v in row_values if v is not None)
        print(f"  Fila {row_num}: {non_empty} celdas con datos")
        if non_empty > 0:
            header_row = row_num
            break

    # Encabezados
    print(f"\n ENCABEZADOS (FILA {header_row}):")
    print(f"{'No.':<5} {'Columna':<5} {'Nombre':<60}")
    print("-" * 100)

    headers = []
    for i in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=i)
        header_name = str(cell.value).strip() if cell.value else f"[Vacio Col{i}]"
        headers.append(header_name)
        col_letter = cell.column_letter
        print(f"{i:<5} {col_letter:<5} {header_name:<60}")

    # Primeras filas de datos
    start_data_row = header_row + 1
    print(f"\n PRIMERAS 5 FILAS DE DATOS (desde fila {start_data_row}):\n")
    for row_num in range(start_data_row, min(start_data_row + 5, ws.max_row + 1)):
        print(f"Fila {row_num}:")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            valor = cell.value
            if valor is None:
                valor = "[Vacio]"
            print(f"  {header:<40}: {valor}")
        print()

    print(f"{'='*100}")
    print("\n Usa 'Fila %d' como header_row en el import_docentes.py\n" % header_row)

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
