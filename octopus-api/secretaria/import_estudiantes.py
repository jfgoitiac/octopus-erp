"""
Parser de la planilla de matrícula histórica (.xlsx) para cargar el Banco de
Alumnos en bloque. Ver auditoría 2026-07-15: el archivo fuente entregado por
el colegio trae errores de captura (columnas desplazadas, cédulas mal
formateadas por Excel, fechas de nacimiento ausentes) — este módulo los
tolera y los reporta en vez de fallar la importación completa.
"""
import re
from datetime import date, datetime

import openpyxl

# Mapa de nombre de sección (fila divisoria "MATRÍCULA ...") -> grado_seccion
# a persistir. Las salas de preescolar se nombran por edad, igual que
# secretaria/seeds.py y PromocionAlumnosView.MAPA_GRADOS.
MAPA_SECCIONES = {
    'SALA A':        'Sala 3',
    'SALA B':        'Sala 4',
    'SALA C':        'Sala 5',
    '1ER GRADO':     '1er Grado',
    '2DO GRADO':     '2do Grado',
    '3ER GRADO':     '3er Grado',
    '4TO GRADO':     '4to Grado',
    '5TO GRADO':     '5to Grado',
    '6TO GRADO':     '6to Grado',
    '1ER AÑO':       '1er Año',
    '2DO AÑO':       '2do Año',
    '3ER AÑO':       '3er Año',
    '4TO AÑO':       '4to Año',
}

# Índices de columna (0-based) según el encabezado real de la planilla.
COL_APELLIDOS_EST   = 2
COL_NOMBRES_EST      = 3
COL_FECHA_NAC        = 4
COL_SEXO             = 5
COL_CEDULA_EST       = 6
COL_LUGAR_NAC        = 7
COL_DIRECCION        = 9
COL_APELLIDOS_REP    = 10
COL_NOMBRES_REP      = 11
COL_CEDULA_REP       = 12
COL_CORREO_REP       = 13
COL_TELEFONO_REP     = 14

# Secciones con el corrimiento conocido: el teléfono del representante quedó
# capturado en la columna de correo y la de teléfono vino vacía.
SECCIONES_TELEFONO_DESPLAZADO = {'6TO GRADO'}

_CEDULA_LIMPIA_RE = re.compile(r'\D')


_MARCADOR_SECCION_RE = re.compile(r'^MATR[ÍI]CULA\s+(.+)$')


def _texto_seccion(fila):
    """
    Fila divisoria (ej. "MATRÍCULA SALA A"): el texto vive en la celda ancla
    de una celda combinada, cuya columna varía según la fila (normalmente B,
    a veces C), así que se busca en toda la fila en vez de fijar un índice.
    Devuelve el nombre de sección normalizado o None si la fila no es un
    divisor de sección.
    """
    for valor in fila:
        if isinstance(valor, str):
            m = _MARCADOR_SECCION_RE.match(valor.strip().upper())
            if m:
                return m.group(1).strip()
    return None


def _es_fila_vacia(fila):
    return not fila[COL_APELLIDOS_EST] and not fila[COL_NOMBRES_EST]


def _limpiar_texto(valor):
    if valor is None:
        return ''
    return str(valor).strip()


def _normalizar_cedula(valor, validar_longitud=False):
    """
    Devuelve (cedula_normalizada, dudosa). `dudosa=True` cuando el valor
    llegó corrupto (Excel lo interpretó como decimal, ej. 36.111535, y el
    resultado ya perdió dígitos — no es recuperable) o, si `validar_longitud`
    está activo (cédulas de identidad reales, no el ID escolar interno que
    no sigue ese formato), cuando el largo no es plausible para una cédula
    venezolana.
    """
    if valor is None or valor == '':
        return '', False

    if isinstance(valor, float):
        # Excel guardó "36.111.535" como número y solo sobrevivió un punto:
        # el resultado (36.111535) ya perdió dígitos — no es recuperable.
        if not valor.is_integer():
            return '', True
        valor = int(valor)

    digitos = _CEDULA_LIMPIA_RE.sub('', str(valor))
    if validar_longitud and (len(digitos) < 6 or len(digitos) > 9):
        return digitos, True
    return digitos, False


def _normalizar_fecha(valor):
    if valor is None or valor == '':
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        valor = valor.strip()
        for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(valor, fmt).date()
            except ValueError:
                continue
    return None


def _normalizar_genero(valor):
    valor = _limpiar_texto(valor).upper()
    return 'femenino' if valor.startswith('F') else 'masculino'


def parsear_planilla(archivo):
    """
    Lee el .xlsx y devuelve una lista de dicts, uno por estudiante:
    {
        fila_excel, grado_seccion, nombre, apellido, cedula_escolar,
        fecha_nacimiento, genero, lugar_nacimiento, direccion,
        representante: {cedula, nombre, apellido, correo, telefono},
        warnings: [str, ...],
    }
    Filas de sección desconocida o completamente vacías se ignoran.
    """
    wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    resultados = []
    seccion_actual = None
    cedulas_vistas = set()

    for idx, fila_tuple in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        fila = list(fila_tuple) + [''] * (15 - len(fila_tuple))
        fila = [v if v is not None else '' for v in fila]

        nombre_seccion = _texto_seccion(fila)
        if nombre_seccion is not None:
            seccion_actual = nombre_seccion if nombre_seccion in MAPA_SECCIONES else None
            continue

        if _es_fila_vacia(fila) or seccion_actual is None:
            continue

        warnings = []
        errors = []

        cedula_est, cedula_dudosa = _normalizar_cedula(fila[COL_CEDULA_EST])
        if cedula_dudosa:
            warnings.append('cedula_estudiante_dudosa')
        elif cedula_est and cedula_est in cedulas_vistas:
            warnings.append('cedula_estudiante_duplicada')
            cedula_est = ''
        elif cedula_est:
            cedulas_vistas.add(cedula_est)

        correo_rep = _limpiar_texto(fila[COL_CORREO_REP])
        telefono_rep = _limpiar_texto(fila[COL_TELEFONO_REP])
        if seccion_actual in SECCIONES_TELEFONO_DESPLAZADO and not telefono_rep and correo_rep and '@' not in correo_rep:
            telefono_rep, correo_rep = correo_rep, ''
            warnings.append('telefono_representante_recuperado_de_correo')

        cedula_rep, cedula_rep_dudosa = _normalizar_cedula(fila[COL_CEDULA_REP], validar_longitud=True)
        if cedula_rep_dudosa:
            warnings.append('cedula_representante_dudosa')
        if not cedula_rep:
            # Representante.cedula es obligatoria y única — sin ella no se
            # puede crear ni vincular al representante, la fila es inválida.
            errors.append('cedula_representante_faltante')

        fecha_nac = _normalizar_fecha(fila[COL_FECHA_NAC])
        if fecha_nac is None:
            warnings.append('fecha_nacimiento_faltante')

        resultados.append({
            'fila_excel':        idx,
            'grado_seccion':     MAPA_SECCIONES[seccion_actual],
            'nombre':            _limpiar_texto(fila[COL_NOMBRES_EST]),
            'apellido':          _limpiar_texto(fila[COL_APELLIDOS_EST]),
            'cedula_escolar':    cedula_est or None,
            'fecha_nacimiento':  fecha_nac,
            'genero':            _normalizar_genero(fila[COL_SEXO]),
            'lugar_nacimiento':  _limpiar_texto(fila[COL_LUGAR_NAC]),
            'direccion':         _limpiar_texto(fila[COL_DIRECCION]),
            'representante': {
                'cedula':   cedula_rep or None,
                'nombre':   _limpiar_texto(fila[COL_NOMBRES_REP]),
                'apellido': _limpiar_texto(fila[COL_APELLIDOS_REP]),
                'correo':   correo_rep,
                'telefono': telefono_rep,
            },
            'warnings': warnings,
            'errors':   errors,
        })

    wb.close()
    return resultados
