from datetime import date, timedelta
from django.shortcuts import get_object_or_404
from django.utils import timezone


def _cedula_visible(cedula):
    """Devuelve la cédula real o cadena vacía si es una cédula temporal generada (99...)."""
    if cedula and cedula.startswith('99') and len(cedula) >= 18:
        return ''
    return cedula or ''
from django.db import transaction
from django.db import models
from django.db.models import F
from rest_framework.response import Response
from rest_framework import viewsets, status, permissions, generics, mixins
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.views import APIView
from .models import (
    Alumno, Beca, BienNacional, ConfiguracionGrado, ConfiguracionSistema,
    Inscripcion, Representante
)
from .services import generate_temporary_cedula_escolar, dia_limite_pago_global
from .import_estudiantes import parsear_planilla
from authentication.views import IsSystemAdminOrDirector
from usuarios.models import LogAuditoria
from django.db.models import Count
from config.pagination import StandardResultsPagination
from .serializers import (
    AlumnoRetirarSerializer, AlumnoSerializer, AlumnoUpdateSerializer,
    AsignarGradoSerializer, BecaSerializer, BienNacionalSerializer, ConfiguracionGradoSerializer,
    ConfiguracionSistemaSerializer, InscripcionListSerializer, InscripcionSerializer,
    RepresentanteSerializer, RepresentanteCRUDSerializer, LogAuditoriaSerializer,
)


# ─────────────────────────────────────────────
# PERMISO PERSONALIZADO PARA DOCENTES
# ─────────────────────────────────────────────
class IsSecretariaOrAbove(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        
        # El superusuario siempre tiene acceso total por diseño
        if request.user.is_superuser:
            return True
            
        try:
            return (
                request.user.perfil.esta_activo and
                request.user.perfil.rol in ['director', 'sistemas', 'administrador', 'secretaria']
            )
        except Exception:
            return False


class IsDocenteOrAbove(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False

        if request.user.is_superuser:
            return True

        try:
            return (
                request.user.perfil.esta_activo and
                request.user.perfil.rol in ['director', 'sistemas', 'administrador', 'secretaria', 'docente', 'cobranza']
            )
        except Exception:
            return False


class IsFinanzasOrAbove(permissions.BasePermission):
    """Roles con acceso a finanzas (sin sistemas, que no debe ver montos)."""
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False

        if request.user.is_superuser:
            return True

        try:
            return (
                request.user.perfil.esta_activo and
                request.user.perfil.rol in ['director', 'administrador', 'cobranza']
            )
        except Exception:
            return False


class IsSecretariaOrCobranzaOrAbove(permissions.BasePermission):
    """Igual que IsSecretariaOrAbove pero también permite a cobranza y docente
    editar los datos del alumno (incluyendo campos financieros, ver update_info)."""
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False

        if request.user.is_superuser:
            return True

        try:
            return (
                request.user.perfil.esta_activo and
                request.user.perfil.rol in ['director', 'sistemas', 'administrador', 'secretaria', 'cobranza', 'docente']
            )
        except Exception:
            return False


# ─────────────────────────────────────────────
# CONFIGURACIÓN DEL SISTEMA (NUEVO)
# ─────────────────────────────────────────────
class ConfiguracionSistemaView(APIView):
    def get_permissions(self):
        # Lectura (necesaria para el wizard de inscripciones, entre otros): docente o superior.
        # Escritura (abre/cierra período, genera cuotas masivas, etc.): solo admin/director/sistemas.
        if self.request.method == 'GET':
            return [IsDocenteOrAbove()]
        return [IsSystemAdminOrDirector()]

    def get(self, request):
        config = ConfiguracionSistema.objects.first()
        if not config:
            return Response({}, status=status.HTTP_200_OK)
        serializer = ConfiguracionSistemaSerializer(config, context={'request': request})
        return Response(serializer.data)

    def post(self, request):
        config = ConfiguracionSistema.objects.first()
        estaba_abierto = config.inscripciones_abiertas if config else False

        if config:
            serializer = ConfiguracionSistemaSerializer(config, data=request.data, partial=True)
        else:
            serializer = ConfiguracionSistemaSerializer(data=request.data)

        serializer.is_valid(raise_exception=True)
        config = serializer.save()

        # DRF's FileField rechaza string vacío como "invalid" (no lo interpreta como
        # borrar), así que el frontend manda un flag aparte para eliminar un logo.
        hubo_borrado = False
        for campo in ('logo_colegio', 'logo_avec', 'favicon'):
            if str(request.data.get(f'{campo}_clear', '')).lower() in ('true', '1'):
                archivo = getattr(config, campo)
                if archivo:
                    archivo.delete(save=False)
                setattr(config, campo, None)
                hubo_borrado = True
        if hubo_borrado:
            config.save()

        # Propagar el día límite de pago a los alumnos: la mora y las
        # notificaciones leen Alumno.dia_limite_pago, así que sin esta
        # sincronización el valor de Configuración no tendría efecto real.
        # Se sincroniza en CADA guardado (no solo al cambiar el valor) para
        # corregir alumnos desviados; el exclude lo hace idempotente y barato.
        alumnos_actualizados = 0
        if config.dia_limite_pago:
            alumnos_actualizados = Alumno.todos.exclude(
                dia_limite_pago=config.dia_limite_pago
            ).update(dia_limite_pago=config.dia_limite_pago)

        # Al abrir el proceso de inscripciones, generar CuotaInscripcion para todos los alumnos activos
        # y CuotaProyectoInversion para sus representantes (una sola vez por representante, no por hijo)
        cuotas_generadas = 0
        proyectos_generados = 0
        if config.inscripciones_abiertas and not estaba_abierto:
            from cobranza.models import CuotaInscripcion, CuotaProyectoInversion, ParametroGlobal
            from cobranza.services import monto_proyecto_inversion_defecto, tipo_cargo_proyecto_inversion
            from decimal import Decimal

            param = ParametroGlobal.objects.filter(clave="MONTO_INSCRIPCION_DEFECTO").first()
            monto = Decimal(param.valor) if param and param.valor else Decimal('50.00')
            periodo = config.periodo_escolar_activo

            alumnos_activos = list(Alumno.objects.filter(activo=True))
            cuotas_nuevas = [
                CuotaInscripcion(
                    alumno=alumno,
                    periodo_escolar=periodo,
                    monto_usd=monto,
                    pagado=False,
                )
                for alumno in alumnos_activos
            ]
            CuotaInscripcion.objects.bulk_create(cuotas_nuevas, ignore_conflicts=True)
            cuotas_generadas = len(cuotas_nuevas)

            monto_proyecto = monto_proyecto_inversion_defecto()
            tipo_proyecto = tipo_cargo_proyecto_inversion()
            representantes_ids = {alumno.representante_id for alumno in alumnos_activos}
            proyectos_nuevos = [
                CuotaProyectoInversion(
                    representante_id=representante_id,
                    periodo_escolar=periodo,
                    tipo_concepto=tipo_proyecto,
                    numero_cuota=1,
                    monto_usd=monto_proyecto,
                    pagado=False,
                )
                for representante_id in representantes_ids
            ]
            CuotaProyectoInversion.objects.bulk_create(proyectos_nuevos, ignore_conflicts=True)
            proyectos_generados = len(proyectos_nuevos)

        LogAuditoria.objects.create(
            usuario=request.user,
            accion="ACTUALIZACION_CONFIGURACION",
            modulo="SISTEMAS",
            detalles={
                "periodo_activo":              config.periodo_escolar_activo,
                "fecha_inicio_inscripciones":  str(config.fecha_inicio_inscripciones),
                "fecha_fin_inscripciones":     str(config.fecha_fin_inscripciones),
                "dia_limite_pago":             config.dia_limite_pago,
                "cuotas_inscripcion_generadas": cuotas_generadas,
                "proyectos_inversion_generados": proyectos_generados,
                "alumnos_dia_limite_actualizados": alumnos_actualizados,
            }
        )

        response_data = ConfiguracionSistemaSerializer(config, context={'request': request}).data
        if cuotas_generadas > 0:
            response_data['cuotas_inscripcion_generadas'] = cuotas_generadas
        if proyectos_generados > 0:
            response_data['proyectos_inversion_generados'] = proyectos_generados
        if alumnos_actualizados > 0:
            response_data['alumnos_dia_limite_actualizados'] = alumnos_actualizados
        return Response(response_data)


# ─────────────────────────────────────────────
# CARGA MANUAL DE CUOTAS DE INSCRIPCIÓN
# ─────────────────────────────────────────────
class CargarCuotasInscripcionView(APIView):
    """Genera las CuotaInscripcion faltantes del período activo para los
    alumnos activos que aún NO tienen grado_seccion asignado (no inscritos).
    La CuotaProyectoInversion (por representante) solo se genera para los
    representantes cuyos hijos activos estén TODOS sin grado asignado.
    Idempotente: bulk_create(ignore_conflicts=True) más el unique_together
    (alumno, periodo_escolar) del modelo impiden la doble carga sin importar
    cuántas veces se ejecute."""
    permission_classes = [IsSystemAdminOrDirector]

    @transaction.atomic
    def post(self, request):
        from decimal import Decimal
        from cobranza.models import CuotaInscripcion, CuotaProyectoInversion, ParametroGlobal
        from cobranza.services import monto_proyecto_inversion_defecto, tipo_cargo_proyecto_inversion

        config = ConfiguracionSistema.objects.first()
        periodo = config.periodo_escolar_activo if config else None
        if not periodo:
            return Response(
                {"error": "No hay período escolar activo configurado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        param = ParametroGlobal.objects.filter(clave="MONTO_INSCRIPCION_DEFECTO").first()
        monto = Decimal(param.valor) if param and param.valor else Decimal('50.00')

        alumnos_activos = list(Alumno.objects.filter(activo=True))
        alumnos_sin_grado = [a for a in alumnos_activos if not (a.grado_seccion or '').strip()]

        existentes = set(
            CuotaInscripcion.objects
            .filter(periodo_escolar=periodo, alumno_id__in=[a.id for a in alumnos_sin_grado])
            .values_list('alumno_id', flat=True)
        )
        faltantes = [a for a in alumnos_sin_grado if a.id not in existentes]

        cuotas_nuevas = [
            CuotaInscripcion(alumno=a, periodo_escolar=periodo, monto_usd=monto, pagado=False)
            for a in faltantes
        ]
        CuotaInscripcion.objects.bulk_create(cuotas_nuevas, ignore_conflicts=True)

        # CuotaProyectoInversion: una por representante (no por alumno), solo para
        # representantes cuyos hijos activos estén TODOS sin grado asignado.
        monto_proyecto = monto_proyecto_inversion_defecto()
        tipo_proyecto = tipo_cargo_proyecto_inversion()
        representantes_sin_grado_ids = {a.representante_id for a in alumnos_sin_grado}
        representantes_con_hijo_inscrito_ids = {
            a.representante_id for a in alumnos_activos if (a.grado_seccion or '').strip()
        }
        representantes_ids = representantes_sin_grado_ids - representantes_con_hijo_inscrito_ids
        representantes_existentes = set(
            CuotaProyectoInversion.objects
            .filter(periodo_escolar=periodo, representante_id__in=representantes_ids)
            .values_list('representante_id', flat=True)
        )
        representantes_faltantes = representantes_ids - representantes_existentes
        proyectos_nuevos = [
            CuotaProyectoInversion(
                representante_id=representante_id,
                periodo_escolar=periodo,
                tipo_concepto=tipo_proyecto,
                numero_cuota=1,
                monto_usd=monto_proyecto,
                pagado=False,
            )
            for representante_id in representantes_faltantes
        ]
        CuotaProyectoInversion.objects.bulk_create(proyectos_nuevos, ignore_conflicts=True)

        LogAuditoria.objects.create(
            usuario=request.user,
            accion="CARGA_CUOTAS_INSCRIPCION",
            modulo="COBRANZA",
            detalles={
                "periodo_escolar": periodo,
                "monto_usd": str(monto),
                "alumnos_sin_grado": len(alumnos_sin_grado),
                "ya_tenian_cuota": len(existentes),
                "cuotas_generadas": len(cuotas_nuevas),
                "monto_proyecto_inversion": str(monto_proyecto),
                "proyectos_inversion_generados": len(proyectos_nuevos),
            }
        )

        return Response({
            "mensaje": (
                f"{len(cuotas_nuevas)} cuota(s) de inscripción y "
                f"{len(proyectos_nuevos)} cuota(s) de Proyecto de Inversión generada(s) "
                f"para alumnos sin grado asignado del período {periodo}."
            ),
            "periodo_escolar": periodo,
            "alumnos_sin_grado": len(alumnos_sin_grado),
            "ya_tenian_cuota": len(existentes),
            "cuotas_generadas": len(cuotas_nuevas),
            "proyectos_inversion_generados": len(proyectos_nuevos),
        })


# ─────────────────────────────────────────────
# PROMOCIÓN AUTOMÁTICA DE ALUMNOS (NUEVO)
# ─────────────────────────────────────────────
class PromocionAlumnosView(APIView):
    permission_classes = [IsSystemAdminOrDirector]

    # Mapa de promoción de grados
    MAPA_GRADOS = {
        'Sala 3':    'Sala 4',
        'Sala 4':    'Sala 5',
        'Sala 5':    '1er Grado',
        '1er Grado': '2do Grado',
        '2do Grado': '3er Grado',
        '3er Grado': '4to Grado',
        '4to Grado': '5to Grado',
        '5to Grado': '6to Grado',
        '6to Grado': '1er Año',  # Transición Primaria -> Media General
        '1er Año':   '2do Año',
        '2do Año':   '3er Año',
        '3er Año':   '4to Año',
        '4to Año':   '5to Año',
    }
    @transaction.atomic
    def post(self, request):
        config = ConfiguracionSistema.objects.first()
        periodo_origen  = config.periodo_escolar_activo if config else "2024-2025"
        periodo_destino = request.data.get('periodo_destino', '2025-2026')

        alumnos_activos = Alumno.objects.filter(activo=True).exclude(grado_seccion=None)
        promovidos      = []
        no_mapeados     = []

        alumnos_a_promover = []
        for alumno in alumnos_activos:
            grado_actual = alumno.grado_seccion or ''
            # Extraer solo el nombre del grado (sin la sección)
            partes = grado_actual.split(' - ')
            nombre_grado = partes[0].strip()
            seccion      = f" - {partes[1].strip()}" if len(partes) > 1 else ''

            nuevo_grado_nombre = self.MAPA_GRADOS.get(nombre_grado)
            if nuevo_grado_nombre:
                alumno.grado_seccion = f"{nuevo_grado_nombre}{seccion}"
                alumnos_a_promover.append(alumno)
                promovidos.append(alumno.id)
            else:
                no_mapeados.append({
                    "alumno_id": alumno.id,
                    "nombre":    f"{alumno.nombre} {alumno.apellido}",
                    "grado":     grado_actual
                })

        # Optimización 1: Promoción en lote (Bulk Update). 
        # Reduce O(N) queries a 1 sola query de actualización masiva.
        if alumnos_a_promover:
            Alumno.objects.bulk_update(alumnos_a_promover, ['grado_seccion'], batch_size=500)

        # Actualizar período activo
        if config:
            config.periodo_escolar_activo = periodo_destino
            config.save(update_fields=['periodo_escolar_activo'])

        # Sincronizar contadores de cupos tras el movimiento masivo de forma precisa
        from django.db.models import Count
        
        # 1. Obtener los conteos actuales de alumnos activos por grado_seccion
        #    Esto incluye todos los grados donde hay alumnos, post-promoción.
        conteos_alumnos_por_grado = Alumno.objects.filter(activo=True) \
                                            .values('grado_seccion') \
                                            .annotate(total=Count('id')) \
                                            .filter(grado_seccion__isnull=False) # Solo contar si grado_seccion no es nulo
        
        # Convertir a un mapa para acceso rápido: {'grado_seccion': count}
        conteos_map = {item['grado_seccion']: item['total'] for item in conteos_alumnos_por_grado}

        # 2. Obtener todas las configuraciones de grado existentes
        all_config_grados = ConfiguracionGrado.objects.all()
        
        # Lista para bulk_update
        config_grados_to_update = []

        # 3. Actualizar cupos_utilizados para cada configuración de grado
        for config_grado in all_config_grados:
            new_cupos_utilizados = conteos_map.get(config_grado.grado_seccion, 0)
            if config_grado.cupos_utilizados != new_cupos_utilizados:
                config_grado.cupos_utilizados = new_cupos_utilizados
                config_grados_to_update.append(config_grado)

        # 4. Realizar la actualización en lote
        if config_grados_to_update:
            ConfiguracionGrado.objects.bulk_update(config_grados_to_update, ['cupos_utilizados'])

        # Generar CuotaInscripcion del nuevo período para los alumnos promovidos
        # (los no_mapeados, ej. egresados, no continúan y no deben generar cuota)
        from cobranza.models import CuotaInscripcion, CuotaProyectoInversion, ParametroGlobal
        from cobranza.services import monto_proyecto_inversion_defecto, tipo_cargo_proyecto_inversion
        from decimal import Decimal

        param = ParametroGlobal.objects.filter(clave="MONTO_INSCRIPCION_DEFECTO").first()
        monto = Decimal(param.valor) if param and param.valor else Decimal('50.00')

        cuotas_nuevas = [
            CuotaInscripcion(
                alumno=alumno,
                periodo_escolar=periodo_destino,
                monto_usd=monto,
                pagado=False,
            )
            for alumno in alumnos_a_promover
        ]
        cuotas_generadas = 0
        if cuotas_nuevas:
            CuotaInscripcion.objects.bulk_create(cuotas_nuevas, ignore_conflicts=True)
            cuotas_generadas = len(cuotas_nuevas)

        # Generar CuotaProyectoInversion del nuevo período por REPRESENTANTE
        # (una sola vez aunque tenga varios hijos promovidos). Antes esta vista
        # solo generaba CuotaInscripcion: los representantes cuyos hijos se
        # promovían por esta vía nunca quedaban con la cuota de proyecto de
        # inversión del nuevo período en ningún lado del sistema.
        monto_proyecto = monto_proyecto_inversion_defecto()
        tipo_proyecto = tipo_cargo_proyecto_inversion()
        representantes_ids = {alumno.representante_id for alumno in alumnos_a_promover}
        representantes_existentes = set(
            CuotaProyectoInversion.objects
            .filter(periodo_escolar=periodo_destino, representante_id__in=representantes_ids)
            .values_list('representante_id', flat=True)
        )
        representantes_faltantes = representantes_ids - representantes_existentes
        proyectos_nuevos = [
            CuotaProyectoInversion(
                representante_id=representante_id,
                periodo_escolar=periodo_destino,
                tipo_concepto=tipo_proyecto,
                numero_cuota=1,
                monto_usd=monto_proyecto,
                pagado=False,
            )
            for representante_id in representantes_faltantes
        ]
        proyectos_generados = 0
        if proyectos_nuevos:
            CuotaProyectoInversion.objects.bulk_create(proyectos_nuevos, ignore_conflicts=True)
            proyectos_generados = len(proyectos_nuevos)

        LogAuditoria.objects.create(
            usuario=request.user,
            accion="PROMOCION_ALUMNOS",
            modulo="SISTEMAS",
            detalles={
                "total_promovidos":       len(promovidos),
                "periodo_origen":         periodo_origen,
                "periodo_destino":        periodo_destino,
                "alumnos_ids":            promovidos,
                "no_mapeados":            len(no_mapeados),
                "cuotas_inscripcion_generadas": cuotas_generadas,
                "proyectos_inversion_generados": proyectos_generados,
            }
        )

        return Response({
            "mensaje":          f"Se promovieron {len(promovidos)} alumnos correctamente.",
            "total_promovidos": len(promovidos),
            "no_mapeados":      no_mapeados,
            "periodo_destino":  periodo_destino,
            "cuotas_inscripcion_generadas": cuotas_generadas,
            "proyectos_inversion_generados": proyectos_generados,
        }, status=status.HTTP_200_OK)


# ─────────────────────────────────────────────
# QUITAR GRADOS A TODOS LOS ALUMNOS (NUEVO)
# Deja a los alumnos activos sin grado_seccion para arrancar el proceso de
# inscripción del nuevo período desde cero (todos "sin inscribir").
# ─────────────────────────────────────────────
class QuitarGradosAlumnosView(APIView):
    permission_classes = [IsSystemAdminOrDirector]

    @transaction.atomic
    def post(self, request):
        alumnos_con_grado = Alumno.objects.filter(activo=True).exclude(grado_seccion__isnull=True).exclude(grado_seccion='')
        total = alumnos_con_grado.count()
        alumnos_ids = list(alumnos_con_grado.values_list('id', flat=True))

        alumnos_con_grado.update(grado_seccion=None)

        # Libera todos los cupos: ya no hay alumnos activos con grado asignado.
        ConfiguracionGrado.objects.filter(cupos_utilizados__gt=0).update(cupos_utilizados=0)

        LogAuditoria.objects.create(
            usuario=request.user,
            accion="QUITAR_GRADOS_MASIVO",
            modulo="SISTEMAS",
            detalles={
                "total_afectados": total,
                "alumnos_ids":     alumnos_ids,
            }
        )

        return Response({
            "mensaje":         f"Se quitó el grado a {total} alumno{'s' if total != 1 else ''}. Ya quedaron sin inscribir.",
            "total_afectados": total,
        }, status=status.HTTP_200_OK)


# ─────────────────────────────────────────────
# BIEN NACIONAL
# ─────────────────────────────────────────────
class BienNacionalViewSet(viewsets.ModelViewSet):
    queryset           = BienNacional.objects.select_related('responsable_asignado').all()
    serializer_class   = BienNacionalSerializer
    permission_classes = [permissions.IsAuthenticated, IsSystemAdminOrDirector]

    def perform_create(self, serializer):
        instance = serializer.save()
        LogAuditoria.objects.create(
            usuario=self.request.user,
            accion="REGISTRO_BIEN_NACIONAL",
            modulo="INVENTARIO",
            detalles={
                "codigo":      instance.codigo_inventario,
                "descripcion": instance.descripcion,
                "ubicacion":   instance.ubicacion,
            }
        )


# ─────────────────────────────────────────────
# ALUMNOS
# ─────────────────────────────────────────────
def _eliminar_alumno_definitivo(alumno, usuario):
    """
    TODO-TEMPORAL: helper de borrado físico, remover junto con las acciones
    eliminar_definitivo (Alumno y Representante) tras la limpieza de datos
    de prueba. Distinto del soft-delete normal (retirar/reactivar): esto
    borra el registro y su historial de forma irreversible.

    Borra primero las relaciones con on_delete=PROTECT (Pago, Inscripcion)
    para que Django no bloquee el borrado; el resto (Mensualidad,
    CuotaInscripcion, CuotaSolvencia, Nota, Asistencia) cae solo por CASCADE
    al borrar el alumno.
    """
    from cobranza.models import Pago

    LogAuditoria.objects.create(
        usuario=usuario,
        accion="ELIMINACION_DEFINITIVA_ALUMNO",
        modulo="SECRETARIA",
        detalles={
            "alumno_id":      alumno.id,
            "nombre":         f"{alumno.nombre} {alumno.apellido}",
            "cedula_escolar": alumno.cedula_escolar,
        }
    )
    Pago.objects.filter(alumno=alumno).delete()
    Inscripcion.objects.filter(alumno=alumno).delete()
    alumno.delete()
class AlumnoListView(viewsets.ModelViewSet):
    serializer_class   = AlumnoSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class    = StandardResultsPagination

    def get_queryset(self):
        # Docentes y secretaria ven solo activos
        # Con ?todos=true los admins pueden ver retirados también
        mostrar_todos = self.request.query_params.get('todos', 'false').lower() == 'true'
        try:
            rol = self.request.user.perfil.rol
        except Exception:
            rol = ''

        if mostrar_todos and rol in ['director', 'administrador', 'sistemas']:
            qs = Alumno.todos.select_related('representante').all()
        else:
            qs = Alumno.objects.select_related('representante').all()

        # Filtro específico por cédula escolar (para validaciones de duplicados)
        cedula = self.request.query_params.get('cedula', '')
        if cedula:
            qs = Alumno.todos.filter(cedula_escolar=cedula)

        # Estatus financiero EN VIVO: se anota `en_mora` con el mismo criterio
        # que el módulo de morosos, para que ambos coincidan sin depender de la
        # tarea Celery. El serializer lee esta anotación en to_representation().
        from cobranza.mora import annotate_en_mora
        qs = annotate_en_mora(qs)

        # Estado de inscripción EN VIVO para el período escolar activo: se anota
        # `inscrito_periodo_activo` para que el serializer no dependa de
        # `grado_seccion` (que no se limpia al cambiar de año escolar).
        config = ConfiguracionSistema.objects.first()
        periodo_activo = config.periodo_escolar_activo if config else None
        if periodo_activo:
            qs = qs.annotate(
                inscrito_periodo_activo=models.Exists(
                    Inscripcion.objects.filter(alumno_id=models.OuterRef('pk'), periodo_escolar=periodo_activo)
                )
            )

        # Prefetch de solvencia (por alumno) y proyecto de inversión (por
        # representante) del período activo: evita que AlumnoSerializer dispare
        # 2 queries por cada uno de sus 4 SerializerMethodField relacionados,
        # por cada alumno de la página (hasta 160 queries extra en una página
        # de 20 sin este prefetch).
        from cobranza.models import CuotaSolvencia, CuotaProyectoInversion
        cuota_solvencia_qs = (
            CuotaSolvencia.objects.filter(periodo_escolar=periodo_activo)
            if periodo_activo else CuotaSolvencia.objects.none()
        )
        cuota_proyecto_qs = (
            CuotaProyectoInversion.objects.filter(periodo_escolar=periodo_activo)
            if periodo_activo else CuotaProyectoInversion.objects.none()
        )
        qs = qs.prefetch_related(
            models.Prefetch(
                'cuotas_solvencia',
                queryset=cuota_solvencia_qs,
                to_attr='_cuota_solvencia_periodo_activo',
            ),
            models.Prefetch(
                'representante__cuotas_proyecto_inversion',
                queryset=cuota_proyecto_qs,
                to_attr='_cuota_proyecto_periodo_activo',
            ),
        )

        # Filtro por estatus financiero (mora, solvente, becado) sobre el estado real
        estatus = self.request.query_params.get('estatus', '')
        if estatus == 'mora':
            qs = qs.filter(en_mora=True).exclude(estatus_financiero='becado')
        elif estatus == 'solvente':
            qs = qs.filter(en_mora=False).exclude(estatus_financiero='becado')
        elif estatus == 'becado':
            qs = qs.filter(estatus_financiero='becado')

        # Búsqueda por nombre, cédula o representante
        buscar = self.request.query_params.get('buscar', '')
        if buscar:
            qs = qs.filter(
                models.Q(nombre__icontains=buscar) |
                models.Q(apellido__icontains=buscar) |
                models.Q(cedula_escolar__icontains=buscar) |
                models.Q(representante__nombre__icontains=buscar) |
                models.Q(representante__cedula__icontains=buscar)
            )
        return qs

    def get_permissions(self):
        # Crear/editar: secretaria o superior
        # Editar info del alumno (update/partial_update/update_info): también cobranza y docente
        # Listar/ver: docente o superior
        # Eliminación definitiva (temporal): solo director/sistemas/admin
        if self.action in ['eliminar_definitivo', 'eliminar_todos']:
            return [permissions.IsAuthenticated(), IsSystemAdminOrDirector()]
        if self.action in ['update', 'partial_update', 'update_info']:
            return [IsSecretariaOrCobranzaOrAbove()]
        if self.action in ['create', 'destroy', 'quitar_grado']:
            return [IsSecretariaOrAbove()]
        return [IsDocenteOrAbove()]

    def perform_create(self, serializer):
        alumno = serializer.save()
        LogAuditoria.objects.create(
            usuario=self.request.user,
            accion="REGISTRO_ALUMNO",
            modulo="SECRETARIA",
            detalles={
                "alumno_id":      alumno.id,
                "nombre":         f"{alumno.nombre} {alumno.apellido}",
                "cedula_escolar": alumno.cedula_escolar,
            }
        )

    @action(detail=True, methods=['patch'])
    @transaction.atomic
    def update_info(self, request, pk=None):
        alumno     = self.get_object()
        serializer = AlumnoUpdateSerializer(alumno, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)

        # El monto/concepto de solvencia es un dato financiero: solo finanzas puede
        # cambiarlo, aunque secretaría pueda editar el resto de los datos del alumno.
        # El frontend siempre manda ambos campos en el payload, así que solo bloqueamos
        # cuando el valor efectivamente difiere del actual (no por su sola presencia).
        cuota_solvencia_actual = None
        if 'monto_solvencia' in serializer.validated_data or 'concepto_solvencia' in serializer.validated_data:
            from cobranza.models import CuotaSolvencia
            config  = ConfiguracionSistema.objects.first()
            periodo = config.periodo_escolar_activo if config else None
            cuota_solvencia_actual = CuotaSolvencia.objects.filter(alumno=alumno, periodo_escolar=periodo).first()

            monto_cambia = 'monto_solvencia' in serializer.validated_data and str(
                cuota_solvencia_actual.monto_usd if cuota_solvencia_actual else None
            ) != str(serializer.validated_data['monto_solvencia'])
            concepto_cambia = 'concepto_solvencia' in serializer.validated_data and str(
                cuota_solvencia_actual.concepto if cuota_solvencia_actual else ''
            ) != str(serializer.validated_data['concepto_solvencia'])

            if (monto_cambia or concepto_cambia) and not IsFinanzasOrAbove().has_permission(request, self):
                return Response(
                    {"error": "No tiene permiso para modificar el monto de solvencia."},
                    status=status.HTTP_403_FORBIDDEN,
                )

        cambios_detectados = []
        for attr, value in serializer.validated_data.items():
            if attr == 'representante':
                rep_instance = alumno.representante
                for rep_attr, rep_value in value.items():
                    valor_actual = getattr(rep_instance, rep_attr, None)
                    if str(valor_actual) != str(rep_value):
                        cambios_detectados.append(f"Rep.{rep_attr}: {valor_actual} -> {rep_value}")
            elif attr == 'monto_solvencia':
                periodo = config.periodo_escolar_activo if config else None
                valor_actual = cuota_solvencia_actual.monto_usd if cuota_solvencia_actual else None
                if str(valor_actual) != str(value):
                    cambios_detectados.append(f"Solvencia {periodo}: {valor_actual} -> {value}")
            else:
                valor_actual = getattr(alumno, attr, None)
                if str(valor_actual) != str(value):
                    cambios_detectados.append(f"Alu.{attr}: {valor_actual} -> {value}")

        serializer.save()
        alumno.refresh_from_db()

        LogAuditoria.objects.create(
            usuario=request.user,
            accion="ACTUALIZACION_DATOS_ALUMNO",
            modulo="SECRETARIA",
            detalles={
                "alumno_id": alumno.id,
                "nombre":    f"{alumno.nombre} {alumno.apellido}",
                "cambios":   cambios_detectados
            }
        )
        # Responder con el estatus en vivo (criterio canónico)
        from cobranza.mora import annotate_en_mora
        alumno_anotado = annotate_en_mora(Alumno.todos.filter(pk=alumno.pk)).first() or alumno
        return Response(AlumnoSerializer(alumno_anotado).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])  # NUEVO
    def asignar_grado(self, request, pk=None):
        """Asigna o cambia el grado de un alumno del banco."""
        alumno     = self.get_object()
        serializer = AsignarGradoSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        grado_nuevo = serializer.validated_data['grado_seccion']

        # Verificar cupos disponibles
        try:
            config = ConfiguracionGrado.objects.get(grado_seccion=grado_nuevo)
            if config.cupos_disponibles <= 0:
                return Response(
                    {"error": f"No hay cupos disponibles en {grado_nuevo}."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except ConfiguracionGrado.DoesNotExist:
            return Response(
                {"error": f"El grado {grado_nuevo} no está configurado en el sistema."},
                status=status.HTTP_400_BAD_REQUEST
            )

        grado_anterior        = alumno.grado_seccion
        alumno.grado_seccion  = grado_nuevo
        alumno.save(update_fields=['grado_seccion'])

        LogAuditoria.objects.create(
            usuario=request.user,
            accion="ASIGNACION_GRADO",
            modulo="SECRETARIA",
            detalles={
                "alumno_id":      alumno.id,
                "nombre":         f"{alumno.nombre} {alumno.apellido}",
                "grado_anterior": grado_anterior,
                "grado_nuevo":    grado_nuevo,
            }
        )
        return Response(
            {"mensaje": f"Grado asignado: {grado_nuevo}"},
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=['post'])  # NUEVO
    @transaction.atomic
    def quitar_grado(self, request, pk=None):
        """Quita el grado asignado a un alumno individual, liberando su cupo."""
        alumno = self.get_object()
        grado_anterior = alumno.grado_seccion
        if not grado_anterior:
            return Response(
                {"error": "El alumno no tiene un grado asignado."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Mismo patrón de locking que retirar()/reactivar(), para decrementar
        # el cupo de forma atómica.
        config = ConfiguracionGrado.objects.select_for_update().filter(
            grado_seccion=grado_anterior
        ).first()
        if config and config.cupos_utilizados > 0:
            ConfiguracionGrado.objects.filter(pk=config.pk).update(
                cupos_utilizados=F('cupos_utilizados') - 1
            )

        alumno.grado_seccion = None
        alumno.save(update_fields=['grado_seccion'])

        LogAuditoria.objects.create(
            usuario=request.user,
            accion="QUITAR_GRADO_INDIVIDUAL",
            modulo="SECRETARIA",
            detalles={
                "alumno_id":      alumno.id,
                "nombre":         f"{alumno.nombre} {alumno.apellido}",
                "grado_anterior": grado_anterior,
            }
        )
        return Response(
            {"mensaje": f"Se quitó el grado {grado_anterior} al alumno."},
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=['post'])  # NUEVO
    def retirar(self, request, pk=None):
        """Soft delete — retira sin eliminar historial."""
        alumno = self.get_object()
        if not alumno.activo:
            return Response(
                {"error": "El alumno ya está retirado."},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = AlumnoRetirarSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        motivo = serializer.validated_data.get('motivo', '')
        alumno.retirar(motivo=motivo)

        LogAuditoria.objects.create(
            usuario=request.user,
            accion="RETIRO_ALUMNO",
            modulo="SECRETARIA",
            detalles={
                "alumno_id":      alumno.id,
                "nombre":         f"{alumno.nombre} {alumno.apellido}",
                "cedula_escolar": alumno.cedula_escolar,
                "motivo":         motivo,
            }
        )
        return Response(
            {"mensaje": f"Alumno {alumno.nombre} {alumno.apellido} retirado."},
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=['post'])  # NUEVO
    def reactivar(self, request, pk=None):
        """Reactiva un alumno retirado."""
        try:
            alumno = Alumno.todos.get(pk=pk)
        except Alumno.DoesNotExist:
            return Response({"error": "Alumno no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        if alumno.activo:
            return Response({"error": "El alumno ya está activo."}, status=status.HTTP_400_BAD_REQUEST)

        alumno.reactivar()

        LogAuditoria.objects.create(
            usuario=request.user,
            accion="REACTIVACION_ALUMNO",
            modulo="SECRETARIA",
            detalles={
                "alumno_id": alumno.id,
                "nombre":    f"{alumno.nombre} {alumno.apellido}",
            }
        )
        return Response(
            {"mensaje": f"Alumno {alumno.nombre} {alumno.apellido} reactivado."},
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=['delete'])  # TODO-TEMPORAL: quitar tras limpieza de datos de prueba
    @transaction.atomic
    def eliminar_definitivo(self, request, pk=None):
        """Borrado físico real (no soft-delete) del alumno. Irreversible."""
        try:
            alumno = Alumno.todos.get(pk=pk)
        except Alumno.DoesNotExist:
            return Response({"error": "Alumno no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        _eliminar_alumno_definitivo(alumno, request.user)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['delete'])  # TODO-TEMPORAL: quitar tras limpieza de datos de prueba
    @transaction.atomic
    def eliminar_todos(self, request):
        """Borrado físico real de TODOS los alumnos. Irreversible."""
        alumnos = list(Alumno.todos.all())
        total = len(alumnos)

        LogAuditoria.objects.create(
            usuario=request.user,
            accion="ELIMINACION_DEFINITIVA_TODOS_ALUMNOS",
            modulo="SECRETARIA",
            detalles={"total": total}
        )
        for alumno in alumnos:
            _eliminar_alumno_definitivo(alumno, request.user)
        return Response({"eliminados": total}, status=status.HTTP_200_OK)

    def perform_destroy(self, instance):
        """Sobreescribir DELETE para usar soft delete."""
        LogAuditoria.objects.create(
            usuario=self.request.user, # Log the user performing the action
            accion="RETIRO_ALUMNO",
            modulo="SECRETARIA",
            detalles={
                "alumno_id":      instance.id,
                "nombre":         f"{instance.nombre} {instance.apellido}",
                "cedula_escolar": instance.cedula_escolar,
                "motivo":         "Eliminación desde interfaz",
            }
        )
        instance.retirar(motivo="Eliminación desde interfaz")


# ─────────────────────────────────────────────
# INSCRIPCIÓN
# ─────────────────────────────────────────────
class InscripcionNuevaView(APIView):
    """Inscribe a un alumno ya registrado en el banco."""
    permission_classes = [permissions.IsAuthenticated]
    serializer_class   = InscripcionSerializer

    def post(self, request):
        # El serializador ahora maneja la atomicidad y la lógica de negocio
        serializer = self.serializer_class(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        inscripcion = serializer.save()
        
        LogAuditoria.objects.create(
            usuario=request.user,
            accion="INSCRIPCION_NUEVO_INGRESO",
            modulo="SECRETARIA",
            detalles={
                "alumno_id":        inscripcion.alumno.id,
                "nombre":           f"{inscripcion.alumno.nombre} {inscripcion.alumno.apellido}",
                "grado_seccion":    inscripcion.grado_seccion,
                "inscripcion_id":   inscripcion.id,
            }
        )

        from notificaciones.tasks import task_notificar_comprobante_inscripcion
        task_notificar_comprobante_inscripcion.delay(inscripcion.id)

        return Response(
            {
                "mensaje":        "Inscripción exitosa",
                "alumno_id":      inscripcion.alumno.id,
                "inscripcion_id": inscripcion.id,
            },
            status=status.HTTP_201_CREATED
        )


# ─────────────────────────────────────────────
# CONSULTA DE INSCRIPCIÓN (para reimpresión de comprobantes)
# ─────────────────────────────────────────────
class InscripcionListView(generics.ListAPIView):
    """Listado de inscripciones por nombre/apellido del alumno, disponible para
    cualquier rol autenticado — se usa para localizar y reimprimir comprobantes."""
    serializer_class    = InscripcionListSerializer
    permission_classes  = [permissions.IsAuthenticated]
    pagination_class    = StandardResultsPagination

    def get_queryset(self):
        qs = Inscripcion.objects.select_related('alumno', 'alumno__representante').order_by('-fecha_inscripcion')
        buscar = self.request.query_params.get('buscar', '')
        if buscar:
            qs = qs.filter(
                models.Q(alumno__nombre__icontains=buscar) |
                models.Q(alumno__apellido__icontains=buscar)
            )
        return qs


# ─────────────────────────────────────────────
# COMPROBANTE DE INSCRIPCIÓN (.docx) — misma planilla oficial que la
# pre-inscripción (utils_preinscripcion.generar_planilla_preinscripcion),
# ya con los datos administrativos confirmados de la inscripción.
# ─────────────────────────────────────────────
class ComprobanteInscripcionView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        from django.http import FileResponse
        from .utils_preinscripcion import generar_planilla_preinscripcion
        try:
            inscripcion = Inscripcion.objects.select_related(
                'alumno', 'alumno__representante', 'usuario_registro'
            ).get(pk=pk)
        except Inscripcion.DoesNotExist:
            return Response({"error": "Inscripción no encontrada."}, status=status.HTTP_404_NOT_FOUND)

        try:
            buffer = generar_planilla_preinscripcion(inscripcion.alumno, inscripcion, campos_seleccionados=None)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        alumno = inscripcion.alumno
        nombre = f"Comprobante_Inscripcion_{alumno.apellido}_{alumno.nombre}_{pk}".replace(' ', '_')
        return FileResponse(
            buffer,
            as_attachment=True,
            filename=f"{nombre}.docx",
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        )


# ─────────────────────────────────────────────
# PLANILLA DE PRE-INSCRIPCIÓN (.docx) — NUEVO
# ─────────────────────────────────────────────
class PlanillaPreinscripcionView(APIView):
    """Rellena la planilla oficial (.docx) para un alumno.

    Se busca por Alumno, no por Inscripcion: al ser una PRE-inscripción, el
    estudiante puede todavía no estar inscrito — si tiene una inscripción
    registrada se usa para el bloque de datos administrativos, pero no es
    un requisito para generar la planilla.
    """
    permission_classes = [IsSecretariaOrAbove]

    def post(self, request, pk):
        from django.http import FileResponse
        from .utils_preinscripcion import generar_planilla_preinscripcion, nombre_archivo_alumno
        try:
            alumno = Alumno.objects.select_related('representante').get(pk=pk)
        except Alumno.DoesNotExist:
            return Response({"error": "Alumno no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        inscripcion = Inscripcion.objects.filter(alumno=alumno).order_by('-fecha_inscripcion').first()
        campos = request.data.get('campos') or []
        try:
            buffer = generar_planilla_preinscripcion(alumno, inscripcion, campos)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return FileResponse(
            buffer,
            as_attachment=True,
            filename=nombre_archivo_alumno(alumno),
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        )


class PlanillaPreinscripcionMasivaView(APIView):
    """Genera la planilla de pre-inscripción para todos los alumnos activos.

    `formato` en el body: 'individual' (default, comportamiento histórico —
    un .zip con un .docx por alumno) o 'unico' (un solo .docx con todas las
    planillas, cada una en página nueva).
    """
    permission_classes = [IsSecretariaOrAbove]

    def post(self, request):
        from django.http import FileResponse
        from .utils_preinscripcion import generar_documento_unico_preinscripciones, generar_zip_preinscripciones

        campos = request.data.get('campos') or []
        formato = request.data.get('formato') or 'individual'
        alumnos = Alumno.objects.filter(activo=True).select_related('representante').order_by('apellido', 'nombre')

        if formato == 'unico':
            try:
                docx_buffer = generar_documento_unico_preinscripciones(alumnos, campos)
            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            return FileResponse(
                docx_buffer,
                as_attachment=True,
                filename=f"PreInscripciones_{date.today().strftime('%Y%m%d')}.docx",
                content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            )

        try:
            zip_buffer = generar_zip_preinscripciones(alumnos, campos)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return FileResponse(
            zip_buffer,
            as_attachment=True,
            filename=f"PreInscripciones_{date.today().strftime('%Y%m%d')}.zip",
            content_type='application/zip',
        )


# ─────────────────────────────────────────────
# EXPORTAR ALUMNOS A EXCEL
# ─────────────────────────────────────────────
class ExportarAlumnosExcelView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from cobranza.exports import ExcelExporter
        from django.db.models import Q as DQ

        from cobranza.mora import annotate_en_mora, estatus_financiero_actual

        estatus = request.query_params.get('estatus', '')
        buscar  = request.query_params.get('buscar', '')

        # Estatus en vivo (mismo criterio que morosos) para que la exportación
        # coincida con lo que se ve en pantalla.
        qs = annotate_en_mora(
            Alumno.objects.filter(activo=True).select_related('representante').order_by('apellido', 'nombre')
        )
        if estatus == 'mora':
            qs = qs.filter(en_mora=True).exclude(estatus_financiero='becado')
        elif estatus == 'solvente':
            qs = qs.filter(en_mora=False).exclude(estatus_financiero='becado')
        elif estatus == 'becado':
            qs = qs.filter(estatus_financiero='becado')
        if buscar:
            qs = qs.filter(
                DQ(nombre__icontains=buscar) |
                DQ(apellido__icontains=buscar) |
                DQ(cedula_escolar__icontains=buscar)
            )

        columns = [
            ('Nombre',          'nombre'),
            ('Apellido',        'apellido'),
            ('Cédula Escolar',  'cedula_escolar'),
            ('Grado / Sección', 'grado_seccion'),
            ('Género',          'genero'),
            ('Estatus',         estatus_financiero_actual),
            ('Representante',   lambda x: f"{x.representante.nombre} {x.representante.apellido}" if x.representante else ''),
            ('Tel. Rep.',       lambda x: x.representante.telefono if x.representante else ''),
        ]

        return ExcelExporter.export(qs, columns, 'lista_alumnos')


class ExportarRepresentantesExcelView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from cobranza.exports import ExcelExporter
        from django.db.models import Q as DQ

        buscar    = request.query_params.get('buscar', '').strip()
        min_hijos = request.query_params.get('min_hijos', '')

        qs = Representante.objects.filter(activo=True).annotate(
            cantidad_alumnos=Count('alumnos', filter=models.Q(alumnos__activo=True))
        ).order_by('apellido', 'nombre')

        if buscar:
            qs = qs.filter(
                DQ(cedula__icontains=buscar)   |
                DQ(nombre__icontains=buscar)   |
                DQ(apellido__icontains=buscar) |
                DQ(correo__icontains=buscar)
            )
        if min_hijos:
            try:
                qs = qs.filter(cantidad_alumnos__gte=int(min_hijos))
            except ValueError:
                pass

        columns = [
            ('Cédula',          'cedula'),
            ('Nombre',          'nombre'),
            ('Apellido',        'apellido'),
            ('Teléfono',        'telefono'),
            ('Correo',          'correo'),
            ('Dirección',       'direccion'),
            ('Alumnos activos', lambda x: x.cantidad_alumnos),
        ]

        return ExcelExporter.export(qs, columns, 'lista_representantes')


# ─────────────────────────────────────────────
# REPRESENTANTE
# ─────────────────────────────────────────────
@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def buscar_representante_por_cedula(request, cedula):
    try:
        representante = Representante.objects.get(cedula=cedula)
        return Response({
            "id":        representante.id,
            "nombre":    representante.nombre,
            "apellido":  representante.apellido,
            "cedula":    representante.cedula,
            "telefono":  representante.telefono,
            "correo":    representante.correo,
            "direccion": representante.direccion,
            "existe":    True
        }, status=status.HTTP_200_OK)
    except Representante.DoesNotExist:
        return Response({"existe": False}, status=status.HTTP_200_OK)

class RepresentanteAlumnosView(APIView):
    """
    Busca un representante por cédula y devuelve su información junto 
    con la lista de alumnos asociados.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, cedula):
        try:
            from cobranza.mora import annotate_en_mora
            representante = Representante.objects.get(cedula=cedula)
            # Estatus en vivo (criterio canónico) — el serializer lo lee de la anotación.
            # select_related evita que AlumnoSerializer dispare una query aparte por
            # cada alumno para anidar RepresentanteSerializer(instance.representante).
            alumnos = annotate_en_mora(
                Alumno.objects.filter(representante=representante).select_related('representante')
            )

            return Response({
                "representante": RepresentanteSerializer(representante).data,
                "alumnos": AlumnoSerializer(alumnos, many=True).data
            }, status=status.HTTP_200_OK)
        except Representante.DoesNotExist:
            return Response(
                {"error": "No se encontró el representante en la base de datos."}, 
                status=status.HTTP_404_NOT_FOUND
            )

# ─────────────────────────────────────────────
# CONFIGURACIÓN DE GRADOS
# ─────────────────────────────────────────────
class ConfiguracionGradoViewSet(viewsets.ModelViewSet):
    queryset           = ConfiguracionGrado.objects.all()
    serializer_class   = ConfiguracionGradoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_update(self, serializer):
        instance = serializer.save()
        LogAuditoria.objects.create(
            usuario=self.request.user,
            accion="AJUSTE_CUPOS",
            modulo="SECRETARIA",
            detalles={
                "grado_seccion":  instance.grado_seccion,
                "cupos_maximos":  instance.cupos_maximos,
                "cupos_actuales": instance.cupos_utilizados,
            }
        )


# ─────────────────────────────────────────────
# IMPORTACIÓN MASIVA DE ESTUDIANTES (Banco de Alumnos)
# ─────────────────────────────────────────────
class ImportarEstudiantesView(APIView):
    """Carga masiva de la planilla histórica de matrícula (.xlsx).

    Solo crea/vincula Representante y Alumno (Banco de Alumnos) con su
    grado_seccion — a propósito NO crea Inscripcion ni cuotas, para no
    tratar a estudiantes ya matriculados como una inscripción nueva (ver
    auditoría 2026-07-15). El período de inscripciones se gestiona aparte
    con el flujo normal (CargarCuotasInscripcionView, etc.).

    ?preview=true: parsea el archivo y devuelve el resumen SIN escribir nada.
    Sin ese parámetro: ejecuta la importación dentro de una transacción.
    """
    permission_classes = [permissions.IsAuthenticated, IsSystemAdminOrDirector]

    # Margen que se suma a cupos_maximos cuando la cantidad real de
    # alumnos importados supera el cupo configurado (decisión del usuario:
    # auto-ajustar en vez de bloquear el import).
    MARGEN_CUPOS = 5

    def post(self, request):
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({"error": "Debe adjuntar un archivo .xlsx"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            filas = parsear_planilla(archivo)
        except Exception:
            return Response(
                {"error": "No se pudo leer el archivo. Verifique que sea un .xlsx válido con el formato esperado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not filas:
            return Response(
                {"error": "No se encontraron estudiantes reconocibles en el archivo."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        es_preview = request.query_params.get('preview', '').lower() == 'true'

        if es_preview:
            resumen = [
                {
                    "fila_excel":     f['fila_excel'],
                    "nombre":         f"{f['nombre']} {f['apellido']}".strip(),
                    "grado_seccion":  f['grado_seccion'],
                    "representante":  f"{f['representante']['nombre']} {f['representante']['apellido']}".strip(),
                    "warnings":       f['warnings'],
                    "errors":         f['errors'],
                }
                for f in filas
            ]
            return Response({
                "total":        len(filas),
                "con_errores":  sum(1 for f in filas if f['errors']),
                "con_warnings": sum(1 for f in filas if f['warnings'] and not f['errors']),
                "filas":        resumen,
            })

        return self._confirmar_importacion(request, filas)

    @transaction.atomic
    def _confirmar_importacion(self, request, filas):
        creados = []
        vinculados_representante_existente = []
        omitidos = []
        cupos_ajustados = {}

        for f in filas:
            if f['errors']:
                omitidos.append({"fila_excel": f['fila_excel'], "errores": f['errors']})
                continue

            rep_data = f['representante']
            representante, rep_creado = Representante.objects.get_or_create(
                cedula=rep_data['cedula'],
                defaults={
                    'nombre':    rep_data['nombre'],
                    'apellido':  rep_data['apellido'],
                    'telefono':  rep_data['telefono'],
                    'correo':    rep_data['correo'],
                    'direccion': '',
                }
            )
            if not rep_creado:
                vinculados_representante_existente.append(representante.cedula)
            if not representante.activo:
                representante.activo = True
                representante.fecha_eliminacion = None
                representante.save(update_fields=['activo', 'fecha_eliminacion'])

            cedula_escolar = f['cedula_escolar'] or generate_temporary_cedula_escolar(request.user)

            grado_seccion = f['grado_seccion']
            config_grado, _ = ConfiguracionGrado.objects.select_for_update().get_or_create(
                grado_seccion=grado_seccion,
                defaults={'cupos_maximos': self.MARGEN_CUPOS}
            )

            alumno = Alumno.objects.create(
                cedula_escolar=cedula_escolar,
                nombre=f['nombre'],
                apellido=f['apellido'],
                fecha_nacimiento=f['fecha_nacimiento'],
                genero=f['genero'],
                lugar_nacimiento=f['lugar_nacimiento'],
                direccion=f['direccion'],
                grado_seccion=grado_seccion,
                representante=representante,
                dia_limite_pago=dia_limite_pago_global(),
            )
            creados.append({"id": alumno.id, "nombre": f"{alumno.nombre} {alumno.apellido}", "grado_seccion": grado_seccion})
            cupos_ajustados[grado_seccion] = cupos_ajustados.get(grado_seccion, 0) + 1

        # Sincroniza cupos_utilizados y sube cupos_maximos si el import superó
        # el cupo configurado (decisión del usuario: auto-ajustar, no bloquear).
        for grado_seccion, cantidad_nueva in cupos_ajustados.items():
            config_grado = ConfiguracionGrado.objects.select_for_update().get(grado_seccion=grado_seccion)
            nuevos_cupos_utilizados = config_grado.cupos_utilizados + cantidad_nueva
            config_grado.cupos_utilizados = nuevos_cupos_utilizados
            config_grado.cupos_maximos = max(config_grado.cupos_maximos, nuevos_cupos_utilizados + self.MARGEN_CUPOS)
            config_grado.save(update_fields=['cupos_utilizados', 'cupos_maximos'])

        LogAuditoria.objects.create(
            usuario=request.user,
            accion="IMPORTAR_ESTUDIANTES",
            modulo="SECRETARIA",
            detalles={
                "creados":       len(creados),
                "omitidos":      len(omitidos),
                "grados_afectados": list(cupos_ajustados.keys()),
            }
        )

        return Response({
            "creados":                              len(creados),
            "vinculados_representante_existente":   len(set(vinculados_representante_existente)),
            "omitidos":                              omitidos,
        }, status=status.HTTP_201_CREATED)


# ─────────────────────────────────────────────
# AUDITORÍA
# ─────────────────────────────────────────────
class LogAuditoriaListView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsSystemAdminOrDirector]

    def get(self, request):
        from django.db.models import Q
        from django.utils.dateparse import parse_date
        logs = LogAuditoria.objects.select_related('usuario').all().order_by('-id')

        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        modulo = request.query_params.get('modulo')
        busqueda = request.query_params.get('q')
        if fecha_inicio:
            logs = logs.filter(fecha_hora__date__gte=parse_date(fecha_inicio))
        if fecha_fin:
            logs = logs.filter(fecha_hora__date__lte=parse_date(fecha_fin))
        if modulo and modulo.upper() != 'TODOS':
            logs = logs.filter(modulo__iexact=modulo)
        if busqueda:
            logs = logs.filter(
                Q(accion__icontains=busqueda) |
                Q(usuario__username__icontains=busqueda) |
                Q(usuario__first_name__icontains=busqueda) |
                Q(usuario__last_name__icontains=busqueda)
            )

        total = logs.count()

        # page_size solo limita cuántos registros viajan por página (para no
        # saturar el navegador); el total de registros disponibles (arriba)
        # y la cantidad de páginas nunca se limitan artificialmente.
        try:
            page = max(1, int(request.query_params.get('page', 1)))
            page_size = min(200, max(1, int(request.query_params.get('page_size', 25))))
        except (ValueError, TypeError):
            page, page_size = 1, 25

        offset = (page - 1) * page_size
        logs_pagina = logs[offset:offset + page_size]

        return Response({
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': max(1, (total + page_size - 1) // page_size),
            'results': LogAuditoriaSerializer(logs_pagina, many=True).data,
        })


class ExportarAuditoriaLogExcelView(APIView):
    """Exporta el log de auditoría (LogAuditoria) a Excel, filtrado por rango de fechas y módulo."""
    permission_classes = [permissions.IsAuthenticated, IsSystemAdminOrDirector]

    def get(self, request):
        from django.utils.dateparse import parse_date
        from cobranza.exports import ExcelExporter

        logs = LogAuditoria.objects.select_related('usuario').all().order_by('-fecha_hora')

        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        modulo = request.query_params.get('modulo')
        if fecha_inicio:
            logs = logs.filter(fecha_hora__date__gte=parse_date(fecha_inicio))
        if fecha_fin:
            logs = logs.filter(fecha_hora__date__lte=parse_date(fecha_fin))
        if modulo and modulo.upper() != 'TODOS':
            logs = logs.filter(modulo__iexact=modulo)

        columns = [
            ('Fecha',    lambda x: x.fecha_hora.strftime('%d/%m/%Y %H:%M')),
            ('Usuario',  lambda x: x.usuario.username if x.usuario else 'SISTEMA'),
            ('Acción',   'accion'),
            ('Módulo',   'modulo'),
            ('IP',       lambda x: x.ip_address or ''),
            ('Detalles', lambda x: x.detalles if isinstance(x.detalles, str) else (str(x.detalles) if x.detalles else '')),
        ]

        return ExcelExporter.export(logs, columns, "auditoria_log")


# ─────────────────────────────────────────────
# MÓDULO DE GRADOS / MATRÍCULA POR GRADO
# ─────────────────────────────────────────────
_NOMBRES_GRADO = {
    '1er Grado': 'Primer Grado',
    '2do Grado': 'Segundo Grado',
    '3er Grado': 'Tercer Grado',
    '4to Grado': 'Cuarto Grado',
    '5to Grado': 'Quinto Grado',
    '6to Grado': 'Sexto Grado',
    '1er Año':   'Primer Año',
    '2do Año':   'Segundo Año',
    '3er Año':   'Tercer Año',
    '4to Año':   'Cuarto Año',
    '5to Año':   'Quinto Año',
}

def _nombre_grado_completo(grado_seccion):
    """Devuelve el nombre completo del grado, conservando la sección si existe."""
    partes = grado_seccion.split(' - ', 1)
    nombre = _NOMBRES_GRADO.get(partes[0].strip(), partes[0].strip())
    return f"{nombre} - {partes[1].strip()}" if len(partes) > 1 else nombre


class GradosListView(APIView):
    """Lista todos los grados activos con cantidad de alumnos inscritos."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        grados = (
            Alumno.objects
            .filter(activo=True)
            .exclude(grado_seccion__isnull=True)
            .exclude(grado_seccion='')
            .values('grado_seccion')
            .annotate(total_alumnos=Count('id'))
            .order_by('grado_seccion')
        )
        return Response(list(grados))


class MatriculaGradoView(APIView):
    """Devuelve la lista de alumnos de un grado con orden configurable."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        grado  = request.query_params.get('grado', '').strip()
        orden  = request.query_params.get('orden', 'apellido')  # 'apellido' | 'cedula'

        if not grado:
            return Response({"error": "Debe especificar el parámetro 'grado'."}, status=status.HTTP_400_BAD_REQUEST)

        from cobranza.mora import annotate_en_mora, estatus_financiero_actual
        qs = annotate_en_mora(
            Alumno.objects
            .filter(activo=True, grado_seccion=grado)
            .select_related('representante')
        )

        if orden == 'cedula':
            qs = qs.order_by('cedula_escolar')
        else:
            qs = qs.order_by('apellido', 'nombre')

        data = [
            {
                'id':               a.id,
                'cedula_escolar':   a.cedula_escolar,
                'nombre':           a.nombre,
                'apellido':         a.apellido,
                'genero':           a.genero,
                'grado_seccion':    a.grado_seccion,
                'estatus_financiero': estatus_financiero_actual(a),
                'representante_nombre': f"{a.representante.nombre} {a.representante.apellido}" if a.representante else '',
                'representante_telefono': a.representante.telefono if a.representante else '',
            }
            for a in qs
        ]
        return Response({'grado': grado, 'total': len(data), 'alumnos': data})


class ExportarMatriculaGradoExcelView(APIView):
    """Exporta la matrícula de un grado a Excel."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from cobranza.exports import ExcelExporter

        grado = request.query_params.get('grado', '').strip()
        orden = request.query_params.get('orden', 'apellido')

        if not grado:
            return Response({"error": "Debe especificar el parámetro 'grado'."}, status=status.HTTP_400_BAD_REQUEST)

        qs = Alumno.objects.filter(activo=True, grado_seccion=grado).select_related('representante')
        qs = qs.order_by('cedula_escolar') if orden == 'cedula' else qs.order_by('apellido', 'nombre')

        # Construir manualmente para agregar numeración y encabezado de grado
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        from django.utils import timezone

        nombre_completo = _nombre_grado_completo(grado)

        wb = Workbook()
        ws = wb.active
        ws.title = "Matrícula"

        title_font  = Font(bold=True, size=13)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        center      = Alignment(horizontal="center")

        ws.merge_cells('A1:D1')
        ws['A1'] = f"Matrícula — {nombre_completo}"
        ws['A1'].font      = title_font
        ws['A1'].alignment = center

        ws.merge_cells('A2:D2')
        ws['A2'] = f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].alignment = center

        headers = ['N°', 'Cédula Escolar', 'Nombres', 'Apellidos']
        ws.append([])
        ws.append(headers)
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center

        for idx, alumno in enumerate(qs, start=1):
            ws.append([
                idx,
                _cedula_visible(alumno.cedula_escolar),
                alumno.nombre,
                alumno.apellido,
            ])

        col_widths = [5, 18, 24, 24]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        nombre_archivo = grado.replace(' ', '_').replace('/', '-')
        timestamp = timezone.now().strftime('%Y-%m-%d_%H%M')
        response['Content-Disposition'] = f'attachment; filename="matricula_{nombre_archivo}_{timestamp}.xlsx"'
        wb.save(response)
        return response


class ExportarMatriculaGradoPDFView(APIView):
    """Exporta la matrícula de un grado a PDF con reportlab."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from django.http import HttpResponse
        from django.utils import timezone
        import io

        grado = request.query_params.get('grado', '').strip()
        orden = request.query_params.get('orden', 'apellido')

        if not grado:
            return Response({"error": "Debe especificar el parámetro 'grado'."}, status=status.HTTP_400_BAD_REQUEST)

        qs = Alumno.objects.filter(activo=True, grado_seccion=grado).select_related('representante')
        qs = qs.order_by('cedula_escolar') if orden == 'cedula' else qs.order_by('apellido', 'nombre')

        nombre_completo = _nombre_grado_completo(grado)

        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.5*cm, bottomMargin=1.5*cm, leftMargin=2*cm, rightMargin=2*cm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=14, spaceAfter=4, alignment=TA_CENTER)
        sub_style   = ParagraphStyle('sub',   parent=styles['Normal'], fontSize=9, spaceAfter=12, alignment=TA_CENTER, textColor=colors.HexColor('#666666'))

        primary_color = colors.HexColor('#1E3A5F')

        elements = [
            Paragraph(f"Lista de Matrícula — {nombre_completo}", title_style),
            Paragraph(f"Orden: {'Por Cédula' if orden == 'cedula' else 'Alfabético'} &nbsp;|&nbsp; Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}", sub_style),
            Spacer(1, 0.3*cm),
        ]

        table_data = [['N°', 'Cédula Escolar', 'Nombres', 'Apellidos']]
        for idx, alumno in enumerate(qs, start=1):
            table_data.append([
                str(idx),
                _cedula_visible(alumno.cedula_escolar),
                alumno.nombre,
                alumno.apellido,
            ])

        col_widths = [1.2*cm, 4*cm, 6.5*cm, 6.5*cm]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND',   (0, 0), (-1, 0),  primary_color),
            ('TEXTCOLOR',    (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',     (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1, 0),  9),
            ('ALIGN',        (0, 0), (-1, 0),  'CENTER'),
            ('BOTTOMPADDING',(0, 0), (-1, 0),  7),
            ('TOPPADDING',   (0, 0), (-1, 0),  7),
            # Filas de datos
            ('FONTNAME',     (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE',     (0, 1), (-1, -1), 9),
            ('ALIGN',        (0, 1), (0, -1),  'CENTER'),
            ('ALIGN',        (1, 1), (1, -1),  'CENTER'),
            ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F7FB')]),
            ('TOPPADDING',   (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING',(0, 1), (-1, -1), 5),
            ('GRID',         (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ]))
        elements.append(table)

        # Pie de página con total
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph(f"Total de alumnos: {len(table_data) - 1}", sub_style))

        doc.build(elements)
        buffer.seek(0)

        nombre_archivo = grado.replace(' ', '_').replace('/', '-')
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="matricula_{nombre_archivo}.pdf"'
        return response


# ─────────────────────────────────────────────
# REPRESENTANTES — CRUD COMPLETO
# ─────────────────────────────────────────────
def _eliminar_representante_fisicamente(rep, usuario):
    """
    Borrado físico real (no soft-delete) de un representante, todos sus
    alumnos (activos y retirados, con su historial financiero/académico) y
    su cuenta de acceso al portal si la tiene. Irreversible. Compartido por
    `eliminar_definitivo` (Limpieza de Datos, sin restricciones) y
    `eliminar_definitivo_manual` (módulo Representantes, exige 0 alumnos).
    """
    from cobranza.models import SolvenciaRepresentante

    LogAuditoria.objects.create(
        usuario=usuario,
        accion="ELIMINACION_DEFINITIVA_REPRESENTANTE",
        modulo="SECRETARIA",
        detalles={
            "representante_id": rep.id,
            "nombre":           f"{rep.nombre} {rep.apellido}",
            "cedula":           rep.cedula,
        }
    )

    for alumno in Alumno.todos.filter(representante=rep):
        _eliminar_alumno_definitivo(alumno, usuario)

    # PROTECT: debe borrarse antes que el representante
    SolvenciaRepresentante.objects.filter(representante=rep).delete()

    # Cuenta de acceso al portal (Django User + PerfilUsuario): no cae por
    # CASCADE del lado de RepresentanteUser.representante, hay que borrarla
    # explícitamente para no dejar credenciales huérfanas.
    portal_user = getattr(rep, 'portal_user', None)
    if portal_user is not None:
        portal_user.user.delete()

    rep.delete()  # cascada: CuotaProyectoInversion restante y RepresentanteUser


class RepresentanteViewSet(viewsets.ModelViewSet):
    """CRUD completo de representantes con búsqueda y conteo de alumnos vinculados."""
    serializer_class   = RepresentanteCRUDSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class    = StandardResultsPagination

    def get_queryset(self):
        from cobranza.models import CuotaInscripcion, CuotaProyectoInversion
        from .models import ConfiguracionSistema

        # Exists() anotado para tiene_inscripcion_impaga: antes el serializer
        # disparaba un CuotaInscripcion.objects.filter(...).exists() por cada
        # representante del listado (N+1 real, detectado por
        # RepresentanteViewSetNPlusOneTest) — mismo criterio de anotación que
        # cantidad_alumnos, sin queries extra por fila.
        inscripcion_impaga_qs = CuotaInscripcion.objects.filter(
            alumno__representante=models.OuterRef('pk'), alumno__activo=True, pagado=False,
        )
        qs = Representante.objects.filter(activo=True).select_related('portal_user').annotate(
            cantidad_alumnos=Count('alumnos', filter=models.Q(alumnos__activo=True)),
            # Alumnos retirados vinculados — el frontend lo necesita para
            # decidir si mostrar "Eliminar definitivamente" (exige 0 alumnos,
            # ni activos ni retirados) sin disparar una query extra por fila.
            cantidad_alumnos_retirados=Count('alumnos', filter=models.Q(alumnos__activo=False)),
            _tiene_inscripcion_impaga=models.Exists(inscripcion_impaga_qs),
        )
        # Prefetch de la cuota de Proyecto de Inversión del período activo (a lo
        # sumo una por representante, por unique_together) para que el serializer
        # no dispare una query por fila en el listado.
        config = ConfiguracionSistema.objects.first()
        periodo = config.periodo_escolar_activo if config else None
        # Siempre se adjunta el prefetch (con queryset vacío si no hay período
        # activo) para que el serializer nunca caiga en su fallback por-fila.
        cuota_qs = (
            CuotaProyectoInversion.objects.filter(periodo_escolar=periodo)
            if periodo else CuotaProyectoInversion.objects.none()
        )
        qs = qs.prefetch_related(
            models.Prefetch(
                'cuotas_proyecto_inversion',
                queryset=cuota_qs,
                to_attr='_cuota_proyecto_periodo_activo',
            )
        )
        buscar = self.request.query_params.get('buscar', '').strip()
        if buscar:
            qs = qs.filter(
                models.Q(cedula__icontains=buscar) |
                models.Q(nombre__icontains=buscar)  |
                models.Q(apellido__icontains=buscar) |
                models.Q(correo__icontains=buscar)
            )
        min_hijos = self.request.query_params.get('min_hijos')
        if min_hijos is not None:
            qs = qs.filter(cantidad_alumnos__gte=int(min_hijos))
        return qs.order_by('apellido', 'nombre')

    def get_permissions(self):
        # destroy (soft-delete) y la eliminación definitiva manual comparten
        # permiso con el resto de acciones de cobranza/finanzas sobre
        # representantes (IsFinanzasOrAbove: director, administrador,
        # cobranza) — 'sistemas' queda fuera a propósito, igual que en el
        # resto de finanzas. create/update/partial_update y el
        # eliminar_definitivo histórico (usado por Sistemas → Limpieza de
        # Datos) siguen exigiendo IsSystemAdminOrDirector, sin cambios.
        if self.action in ['destroy', 'eliminar_definitivo_manual']:
            return [permissions.IsAuthenticated(), IsFinanzasOrAbove()]
        if self.action in ['create', 'update', 'partial_update', 'eliminar_definitivo', 'cargar_proyecto_inversion']:
            return [permissions.IsAuthenticated(), IsSystemAdminOrDirector()]
        return [permissions.IsAuthenticated()]

    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        from django.utils import timezone

        rep = self.get_object()

        # Soft delete en cascada: se retiran (no se borran) los alumnos activos
        # vinculados para preservar su historial de pagos/facturas, y el propio
        # representante se marca inactivo en vez de eliminarse físicamente.
        # (Si se hiciera rep.delete() directo, on_delete=CASCADE en
        # Alumno.representante borraría también los alumnos ya retirados.)
        for alumno in rep.alumnos.filter(activo=True):
            alumno.retirar(motivo=f'Representante {rep.cedula} eliminado')

        rep.activo = False
        rep.fecha_eliminacion = timezone.now()
        rep.save(update_fields=['activo', 'fecha_eliminacion'])

        LogAuditoria.objects.create(
            usuario=request.user,
            accion="ELIMINACION_REPRESENTANTE",
            modulo="SECRETARIA",
            detalles={
                "representante_id": rep.id,
                "nombre":           f"{rep.nombre} {rep.apellido}",
                "cedula":           rep.cedula,
            }
        )
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'])
    @transaction.atomic
    def cargar_proyecto_inversion(self, request, pk=None):
        """
        Carga manual (idempotente) de la CuotaProyectoInversion del período
        activo para UN representante puntual, para los casos detectados donde
        el backfill masivo no alcanzó a cubrirlo (representante nuevo, alumno
        reasignado, etc.).

        Restringida a representantes con al menos una CuotaInscripcion sin
        pagar: si ya pagó la inscripción, se asume que el proyecto de
        inversión ya fue cargado por el flujo normal (apertura de
        inscripciones / carga manual / inscripción individual), y forzar la
        carga aquí serviría solo para generar cobros duplicados o fuera de
        lugar sobre cuentas ya al día.
        """
        from cobranza.models import CuotaInscripcion, CuotaProyectoInversion
        from cobranza.services import monto_proyecto_inversion_defecto, tipo_cargo_proyecto_inversion

        rep = self.get_object()
        config = ConfiguracionSistema.objects.first()
        periodo = config.periodo_escolar_activo if config else None
        if not periodo:
            return Response(
                {"error": "No hay un período escolar activo configurado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        tiene_inscripcion_impaga = CuotaInscripcion.objects.filter(
            alumno__representante=rep, alumno__activo=True, pagado=False
        ).exists()
        if not tiene_inscripcion_impaga:
            return Response(
                {"error": (
                    "Este representante no tiene inscripción pendiente; el "
                    "Proyecto de Inversión ya debería estar cargado por el flujo "
                    "normal. Revise la ficha del representante antes de forzar la carga."
                )},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # unique_together=('representante', 'periodo_escolar', 'tipo_concepto',
        # 'numero_cuota') respeta el "un solo cargo por período" incluso si
        # esto se llama más de una vez.
        cuota, creada = CuotaProyectoInversion.objects.get_or_create(
            representante=rep, periodo_escolar=periodo,
            tipo_concepto=tipo_cargo_proyecto_inversion(), numero_cuota=1,
            defaults={'monto_usd': monto_proyecto_inversion_defecto()},
        )

        LogAuditoria.objects.create(
            usuario=request.user,
            accion="CARGA_MANUAL_PROYECTO_INVERSION",
            modulo="SECRETARIA",
            detalles={
                "representante_id": rep.id,
                "cedula":           rep.cedula,
                "periodo_escolar":  periodo,
                "monto_usd":        str(cuota.monto_usd),
                "ya_existia":       not creada,
            }
        )

        return Response({
            "mensaje": (
                f"Proyecto de Inversión del período {periodo} cargado (${cuota.monto_usd})."
                if creada else
                f"Este representante ya tenía el Proyecto de Inversión del período {periodo} cargado."
            ),
            "creada": creada,
            "monto_usd": str(cuota.monto_usd),
            "periodo_escolar": periodo,
        })

    @action(detail=True, methods=['delete'])  # TODO-TEMPORAL: quitar tras limpieza de datos de prueba
    @transaction.atomic
    def eliminar_definitivo(self, request, pk=None):
        """
        Borrado físico real (no soft-delete) del representante, de todos sus
        alumnos (con su historial financiero/académico) y de su cuenta de
        acceso al portal si la tiene. Irreversible.

        Usado hoy solo por Sistemas → Limpieza de Datos, sin restricción de
        "0 alumnos" a propósito: ahí sirve para arrasar datos de prueba con
        alumnos y todo. Para el borrado manual de un representante duplicado
        sin alumnos desde el módulo Representantes, ver `eliminar_definitivo_manual`.
        """
        rep = get_object_or_404(Representante, pk=pk)
        _eliminar_representante_fisicamente(rep, request.user)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['delete'])
    @transaction.atomic
    def eliminar_definitivo_manual(self, request, pk=None):
        """
        Borrado físico real invocado a propósito desde el módulo
        Representantes (no Sistemas → Limpieza de Datos), para casos como un
        representante duplicado por un error de carga de datos: exige que no
        tenga NINGÚN alumno vinculado (ni activo ni retirado) — si tiene
        alguno, se rechaza y debe usarse el retiro/eliminación normal
        (soft-delete, `destroy`) en su lugar. Libera la cédula para reuso.
        """
        rep = get_object_or_404(Representante, pk=pk)

        if Alumno.todos.filter(representante=rep).exists():
            return Response(
                {"error": (
                    "Este representante tiene alumnos vinculados (activos o "
                    "retirados) y no puede eliminarse definitivamente. Use la "
                    "eliminación normal (retira alumnos y preserva el historial)."
                )},
                status=status.HTTP_400_BAD_REQUEST,
            )

        _eliminar_representante_fisicamente(rep, request.user)
        return Response(status=status.HTTP_204_NO_CONTENT)


# ─────────────────────────────────────────────
# BECAS
# ─────────────────────────────────────────────
class BecaViewSet(
    mixins.ListModelMixin, mixins.CreateModelMixin, mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin, viewsets.GenericViewSet,
):
    """
    CRUD de becas, sin destroy físico (ver acción `revocar` — una beca nunca
    se borra, se revoca, para no perder el registro auditable). Solo
    director/administrador/sistemas pueden crear, editar o revocar; el resto
    del staff autenticado puede listar/consultar.
    """
    serializer_class = BecaSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = StandardResultsPagination

    def get_queryset(self):
        qs = Beca.objects.select_related('alumno', 'otorgada_por', 'revocada_por').all()

        alumno_id = self.request.query_params.get('alumno_id')
        if alumno_id:
            qs = qs.filter(alumno_id=alumno_id)

        periodo = self.request.query_params.get('periodo_escolar')
        if periodo:
            qs = qs.filter(periodo_escolar=periodo)

        estado = self.request.query_params.get('estado')
        if estado:
            qs = qs.filter(estado=estado)

        buscar = self.request.query_params.get('buscar', '').strip()
        if buscar:
            qs = qs.filter(
                models.Q(alumno__nombre__icontains=buscar) |
                models.Q(alumno__apellido__icontains=buscar) |
                models.Q(alumno__cedula_escolar__icontains=buscar)
            )

        return qs

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'revocar']:
            return [permissions.IsAuthenticated(), IsSystemAdminOrDirector()]
        return [permissions.IsAuthenticated()]

    def perform_create(self, serializer):
        serializer.save(otorgada_por=self.request.user, estado='activa')

    def perform_update(self, serializer):
        from rest_framework.exceptions import ValidationError as DRFValidationError
        # Editar una beca revocada la "reviviría" fuera del flujo de
        # revocar()/motivo_revocacion — para volver a becar a un alumno se
        # crea una beca nueva, dejando intacto el historial de la revocada.
        if serializer.instance.estado != 'activa':
            raise DRFValidationError('No se puede editar una beca revocada. Cree una nueva.')
        serializer.save()

    @action(detail=True, methods=['post'])
    def revocar(self, request, pk=None):
        from django.utils import timezone

        beca = self.get_object()
        if beca.estado != 'activa':
            return Response({'error': 'La beca ya está revocada.'}, status=status.HTTP_400_BAD_REQUEST)

        beca.estado = 'revocada'
        beca.revocada_por = request.user
        beca.fecha_revocacion = timezone.now()
        beca.motivo_revocacion = request.data.get('motivo', '')
        beca.save(update_fields=['estado', 'revocada_por', 'fecha_revocacion', 'motivo_revocacion'])
        return Response(BecaSerializer(beca).data)


# ─────────────────────────────────────────────
# INDICADOR DE INSCRIPCIONES (Dashboard administrativo)
# ─────────────────────────────────────────────
class InscripcionStatsView(APIView):
    """Indicador de inscripciones para el Dashboard administrativo — mismo
    estilo que cobranza.views.DashboardStatsView (aggregate/annotate a mano,
    sin serializer ni paginación).

    Ver NOTAS_TECNICAS.md: `ConfiguracionGrado` no está sedeada, así que
    `ocupacion.por_grado` queda global aunque el resto del indicador se
    filtre por la sede del usuario, y queda duplicada con la ocupación por
    grado que ya expone `cobranza/stats/` (DashboardStatsView).
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from django.db.models import Count, Min, Max
        from django.db.models.functions import TruncMonth
        from cobranza.permissions import filtrar_por_sede
        from common.periodo import periodo_escolar_activo

        config = ConfiguracionSistema.objects.first()
        periodo_query = request.query_params.get('periodo')

        # Criterio del 400: no hay NINGUNA ConfiguracionSistema cargada Y
        # tampoco se pasó ?periodo= explícito. El fallback hardcodeado del
        # helper compartido (common.periodo.periodo_escolar_activo, usado por
        # otros consumidores como `academico`) no se usa aquí a propósito:
        # este indicador nuevo prefiere fallar con un mensaje claro antes que
        # mostrar cifras de un período inventado ("2025-2026" a ciegas).
        if not periodo_query and config is None:
            return Response(
                {"error": "No hay configuración de período escolar activo. Configure el sistema antes de consultar este indicador."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        periodo = periodo_query or periodo_escolar_activo()

        qs = filtrar_por_sede(
            request.user,
            Inscripcion.objects.filter(periodo_escolar=periodo),
            campo='alumno__sede',
        )

        total_inscritos = qs.count()

        # --- por_tipo_ingreso ---
        conteos_tipo = {
            row['tipo_ingreso']: row['total']
            for row in qs.values('tipo_ingreso').annotate(total=Count('id'))
        }
        por_tipo_ingreso = {
            'nuevo_ingreso': conteos_tipo.get('nuevo', 0),
            'regular':       conteos_tipo.get('regular', 0),
        }

        # --- mes_actual: mes en curso vs mes anterior ---
        ahora = timezone.now()
        anio_actual, mes_actual_num = ahora.year, ahora.month
        if mes_actual_num == 1:
            anio_anterior, mes_anterior_num = anio_actual - 1, 12
        else:
            anio_anterior, mes_anterior_num = anio_actual, mes_actual_num - 1

        cantidad_mes_actual = qs.filter(
            fecha_inscripcion__year=anio_actual, fecha_inscripcion__month=mes_actual_num,
        ).count()
        cantidad_mes_anterior = qs.filter(
            fecha_inscripcion__year=anio_anterior, fecha_inscripcion__month=mes_anterior_num,
        ).count()

        if cantidad_mes_anterior == 0:
            # División por cero indefinida: se reporta `null` (None) en vez de
            # inventar un 0.0 o un 100% — sea porque no hubo inscripciones el
            # mes anterior ni tampoco este (sin variación real que mostrar), o
            # porque hubo inscripciones nuevas partiendo de una base de 0 (el
            # porcentaje de crecimiento no es un número finito). El frontend
            # decide cómo representar `null` (ej. "N/A").
            variacion_pct = None
        else:
            variacion_pct = round(
                (cantidad_mes_actual - cantidad_mes_anterior) / cantidad_mes_anterior * 100, 2
            )

        mes_actual = {
            'mes':           f"{anio_actual:04d}-{mes_actual_num:02d}",
            'cantidad':      cantidad_mes_actual,
            'variacion_pct': variacion_pct,
        }

        # --- documentos_pendientes ---
        documentos_pendientes = qs.filter(documentos_completos=False).count()

        # --- ocupacion ---
        inscritos_por_grado = {
            row['grado_seccion']: row['total']
            for row in qs.values('grado_seccion').annotate(total=Count('id'))
        }
        # ConfiguracionGrado no está sedeada (deuda técnica ya conocida, ver
        # NOTAS_TECNICAS.md): se listan TODAS las configuraciones existentes,
        # cruzadas contra los inscritos del queryset ya filtrado por sede y
        # período. El cruce en sí es correcto; el catálogo de grados que se
        # cruza es global.
        configs_grado = list(ConfiguracionGrado.objects.order_by('grado_seccion'))
        por_grado = []
        for cfg in configs_grado:
            inscritos = inscritos_por_grado.get(cfg.grado_seccion, 0)
            pct = round(inscritos / cfg.cupos_maximos * 100, 2) if cfg.cupos_maximos else 0.0
            por_grado.append({
                'grado_seccion':     cfg.grado_seccion,
                'inscritos':         inscritos,
                'cupos_maximos':     cfg.cupos_maximos,
                'cupos_disponibles': cfg.cupos_disponibles,
                'pct':               pct,
                'sin_cupos':         cfg.cupos_disponibles == 0,
            })

        # global_pct: se decidió como inscritos totales del período / suma de
        # cupos_maximos de TODOS los grados existentes (no solo los que
        # tienen alguna inscripción este período), para que sea consistente
        # con `por_grado` (que también lista todos los grados existentes) y
        # no varíe según qué grados resulten con inscripciones.
        total_cupos_maximos = sum(cfg.cupos_maximos for cfg in configs_grado)
        global_pct = round(total_inscritos / total_cupos_maximos * 100, 2) if total_cupos_maximos else 0.0

        ocupacion = {
            'global_pct': global_pct,
            'por_grado':  por_grado,
        }

        # --- serie_mensual ---
        if config and config.fecha_inicio_ano_escolar and config.fecha_fin_ano_escolar:
            fecha_inicio_rango = config.fecha_inicio_ano_escolar
            fecha_fin_rango = config.fecha_fin_ano_escolar
        else:
            # Solo puede pasar si se pasó ?periodo= a mano sin ninguna
            # ConfiguracionSistema cargada (si no, ya se respondió 400 arriba):
            # no hay fecha_inicio/fin de año escolar de dónde generar el
            # rango. Se usa el rango real de fechas de inscripción del propio
            # queryset filtrado como mejor aproximación disponible; si no hay
            # ninguna inscripción, la serie queda vacía.
            rango = qs.aggregate(inicio=Min('fecha_inscripcion'), fin=Max('fecha_inscripcion'))
            fecha_inicio_rango = rango['inicio'].date() if rango['inicio'] else None
            fecha_fin_rango = rango['fin'].date() if rango['fin'] else None

        serie_mensual = []
        if fecha_inicio_rango and fecha_fin_rango:
            conteos_mes = {
                (row['mes'].year, row['mes'].month): row['cantidad']
                for row in qs.annotate(mes=TruncMonth('fecha_inscripcion')).values('mes').annotate(cantidad=Count('id'))
            }
            y, m = fecha_inicio_rango.year, fecha_inicio_rango.month
            while (y, m) <= (fecha_fin_rango.year, fecha_fin_rango.month):
                serie_mensual.append({
                    'mes':      f"{y:04d}-{m:02d}",
                    'cantidad': conteos_mes.get((y, m), 0),
                })
                m += 1
                if m > 12:
                    m = 1
                    y += 1

        # --- visible: si el bloque de indicadores de inscripciones debe
        # mostrarse HOY en el Dashboard administrativo. Ventana ampliada
        # (-5 días antes de apertura / +15 días después de cierre) respecto
        # a la ventana exacta que usa `ConfiguracionSistema.inscripciones_abiertas`.
        # Si el caller pidió explícitamente un `periodo` por query param
        # (típicamente uno histórico distinto al activo), se ignora la fecha
        # de hoy y siempre se responde visible=True: un admin consultando un
        # período pasado a propósito no debe quedar bloqueado por la ventana.
        if periodo_query:
            visible = True
        else:
            hoy = timezone.now().date()
            ventana_inicio = config.fecha_inicio_inscripciones - timedelta(days=5)
            ventana_fin = config.fecha_fin_inscripciones + timedelta(days=15)
            visible = ventana_inicio <= hoy <= ventana_fin

        return Response({
            'periodo_escolar':       periodo,
            'total_inscritos':       total_inscritos,
            'por_tipo_ingreso':      por_tipo_ingreso,
            'mes_actual':            mes_actual,
            'documentos_pendientes': documentos_pendientes,
            'ocupacion':             ocupacion,
            'serie_mensual':         serie_mensual,
            'visible':               visible,
        })