from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.utils import timezone
from decimal import Decimal, InvalidOperation

from .models import Empleado, TipoCargo, BancoNomina
from .serializers import EmpleadoSerializer, TipoCargoSerializer, BancoNominaSerializer
from cobranza.models import TasaCambio
from authentication.views import IsSystemAdminOrDirector


class TipoCargoViewSet(viewsets.ModelViewSet):
    queryset = TipoCargo.objects.all()
    serializer_class = TipoCargoSerializer
    permission_classes = [IsSystemAdminOrDirector]


class BancoNominaViewSet(viewsets.ModelViewSet):
    queryset = BancoNomina.objects.all()
    serializer_class = BancoNominaSerializer
    permission_classes = [IsSystemAdminOrDirector]

    def get_queryset(self):
        solo_activos = self.request.query_params.get('activos')
        if solo_activos:
            return BancoNomina.objects.filter(activo=True)
        return BancoNomina.objects.all()


class EmpleadoViewSet(viewsets.ModelViewSet):
    queryset = Empleado.objects.all()
    serializer_class = EmpleadoSerializer
    permission_classes = [IsSystemAdminOrDirector]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            return Empleado.objects.none()
        return Empleado.objects.filter(activo=True).select_related('banco')

    @action(detail=False, methods=['get'])
    def preview_bancaribe(self, request):
        from django.db import models as db_models
        empleados = self.get_queryset().filter(
            db_models.Q(numero_cuenta__startswith='0114') |
            db_models.Q(banco__nombre__icontains='bancaribe')
        ).filter(numero_cuenta__gt='', tipo_cuenta__gt='')

        tasa = TasaCambio.objects.first()
        tasa_valor = float(tasa.valor_bs) if tasa else 0

        data = [
            {
                'id': emp.id,
                'nombre': emp.nombre,
                'apellido': emp.apellido,
                'cedula': emp.cedula,
                'numero_cuenta': emp.numero_cuenta,
                'tipo_cuenta': emp.tipo_cuenta,
                'correo': emp.correo or '',
                'telefono': emp.telefono or '',
                'banco_nombre': emp.banco.nombre if emp.banco else '',
            }
            for emp in empleados
        ]
        return Response({'empleados': data, 'tasa': tasa_valor})

    @action(detail=False, methods=['get'])
    def exportar_txt(self, request):
        empleados = self.get_queryset()
        lines = []

        for emp in empleados:
            banco = emp.banco.nombre if emp.banco else ''
            line = (
                f"{emp.cedula};"
                f"{emp.nombre} {emp.apellido};"
                f"{emp.numero_cuenta or ''};"
                f"{banco};"
                f"NOMINA {timezone.now().strftime('%Y-%m')}"
            )
            lines.append(line)

        content = "\n".join(lines)
        response = HttpResponse(content, content_type='text/plain')
        response['Content-Disposition'] = (
            f'attachment; filename="NOMINA_OCTOPUS_'
            f'{timezone.now().strftime("%Y%m%d")}.txt"'
        )
        return response

    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({"error": "openpyxl no está instalado."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Nómina"

        headers = ['Cédula', 'Nombre', 'Apellido', 'Cargo', 'Banco', 'N° Cuenta', 'Teléfono', 'Correo', 'Fecha Ingreso']
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for emp in self.get_queryset():
            ws.append([
                emp.cedula,
                emp.nombre,
                emp.apellido,
                emp.cargo,
                emp.banco.nombre if emp.banco else '',
                emp.numero_cuenta or '',
                emp.telefono or '',
                emp.correo or '',
                emp.fecha_contratacion.strftime('%d/%m/%Y') if emp.fecha_contratacion else '',
            ])

        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="nomina_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        )
        return response

    @action(detail=False, methods=['get'])
    def get_choices(self, request):
        return Response(
            {"cargos": ["Profesor", "Administrativo", "Mantenimiento", "Director"]},
            status=status.HTTP_200_OK
        )
