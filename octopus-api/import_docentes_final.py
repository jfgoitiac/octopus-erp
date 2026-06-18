#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script de importacion de docentes desde Excel a la BD.
Uso: cd C:\Octopus\octopus-api && python import_docentes_final.py <ruta_excel>
     cd C:\Octopus\octopus-api && python import_docentes_final.py
       (usa ruta por defecto del Downloads)
"""
import os
import sys
import django
from openpyxl import load_workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from rrhh.models import Empleado

def limpiar_cedula(cedula):
    if not cedula:
        return None
    cedula = str(cedula).strip().replace('.', '').replace('-', '')
    return cedula if cedula else None

def limpiar_texto(texto):
    if texto is None or texto == '':
        return ''
    return str(texto).strip()

def importar(excel_path):
    if not os.path.exists(excel_path):
        print(f"[ERROR] Archivo no encontrado: {excel_path}")
        return False

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    HEADER_ROW = 5
    headers = {}
    for i in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=HEADER_ROW, column=i).value
        if cell_value:
            headers[cell_value.lower().strip()] = i - 1

    if 'apellido y nombre' not in headers or 'cedula' not in headers:
        print("[ERROR] Columnas requeridas no encontradas (APELLIDO Y NOMBRE, CEDULA)")
        return False

    creados = 0
    actualizados = 0
    errores = 0

    print(f"\nProcesando {ws.max_row - 5} filas...")

    for row_num in range(HEADER_ROW + 1, ws.max_row + 1):
        try:
            nombre_completo = limpiar_texto(
                ws.cell(row=row_num, column=headers['apellido y nombre'] + 1).value
            )
            cedula = limpiar_cedula(
                ws.cell(row=row_num, column=headers['cedula'] + 1).value
            )
            correo = limpiar_texto(
                ws.cell(row=row_num, column=headers.get('correo', -1) + 1).value
                if 'correo' in headers else ''
            )
            telefono = limpiar_texto(
                ws.cell(row=row_num, column=headers.get('telefono', -1) + 1).value
                if 'telefono' in headers else ''
            )

            if not nombre_completo or not cedula:
                continue

            partes = nombre_completo.split()
            if len(partes) >= 2:
                apellido = partes[0]
                nombre = ' '.join(partes[1:])
            else:
                apellido = partes[0]
                nombre = ''

            empleado, creado = Empleado.objects.update_or_create(
                cedula=cedula,
                defaults={
                    'nombre': nombre,
                    'apellido': apellido,
                    'cargo': '',
                    'tipo_personal': 'docente',
                    'titulo': '',
                    'categoria_docente': '',
                    'nivel': '',
                    'horas_semanales': None,
                    'numero_hijos': 0,
                    'telefono': telefono,
                    'correo': correo,
                    'numero_cuenta': '',
                    'tipo_cuenta': '',
                    'sueldo_base': None,
                    'activo': True,
                }
            )

            if creado:
                creados += 1
            else:
                actualizados += 1

        except Exception as e:
            errores += 1
            print(f"[ERROR] Fila {row_num}: {e}")

    print(f"\n[OK] Creados: {creados} | Actualizados: {actualizados} | Errores: {errores}")
    return True

if __name__ == '__main__':
    default_path = r"C:\Users\PC\Downloads\Base de Datos Colegio\DATOS DEL PERSONAL DOCENTE.xlsx"
    path = sys.argv[1] if len(sys.argv) > 1 else default_path

    print(f"Importando docentes desde: {path}")
    success = importar(path)
    sys.exit(0 if success else 1)
