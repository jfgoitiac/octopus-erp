from openpyxl import load_workbook

file_path = r"C:\Users\PC\Downloads\Base de Datos Colegio\DATOS DEL PERSONAL DOCENTE.xlsx"

try:
    wb = load_workbook(file_path)
    print("=== HOJAS DISPONIBLES ===")
    print(wb.sheetnames)

    ws = wb.active
    print(f"\n=== HOJA ACTIVA: {ws.title} ===")

    print("\nEncabezados (Primera fila):")
    headers = []
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers.append(cell.value)
            print(f"  Col {col_idx}: {cell.value}")

    print(f"\nPrimeras 10 filas de datos:")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=False), 2):
        print(f"\nFila {row_idx}:")
        for col_idx, cell in enumerate(row, 1):
            if col_idx <= len(headers):
                print(f"  {headers[col_idx-1]}: {cell.value}")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
