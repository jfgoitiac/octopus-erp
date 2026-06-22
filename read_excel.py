import pandas as pd
from openpyxl import load_workbook

file_path = r"C:\Users\PC\Downloads\Base de Datos Colegio\DATOS DEL PERSONAL DOCENTE.xlsx"

print("=== USANDO OPENPYXL ===")
wb = load_workbook(file_path, data_only=True)
print(f"Hojas: {wb.sheetnames}")
ws = wb.active
print(f"Dimensiones: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

print("\n=== PRIMERAS 20 FILAS Y 10 COLUMNAS ===")
for row_idx in range(1, min(21, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(11, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        row_data.append(str(cell.value)[:20] if cell.value else "")
    print(f"Fila {row_idx}: {row_data}")

print("\n=== USANDO PANDAS ===")
df = pd.read_excel(file_path)
print(f"Shape: {df.shape}")
print(f"\nColumnas: {list(df.columns)}")
print(f"\nPrimeros datos:")
print(df.head())
